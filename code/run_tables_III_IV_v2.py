"""
Table III and Table IV replication — Asness, Moskowitz, Pedersen (2013)

Table III (p.957): Multivariate macro regression
  - Single OLS with ALL macro variables simultaneously
  - HAC standard errors (Newey-West, 6 lags)
  - 4 specifications: US stocks, Global stocks, Nonstock, All

Table IV (p.959-960): Liquidity risk exposures
  - Liquidity shocks = AR(2) residuals of raw liquidity variables
  - Controls: Table III macro variables
  - Dependent variables: Value, Momentum, 50/50 Combo, Val-Mom spread
"""

import numpy as np
import pandas as pd
import statsmodels.api as sm
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / 'data'
OUTPUT_DIR = Path(__file__).parent / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)


# ============================================================
# 1. Factor data
# ============================================================
def load_factors():
    """Load VME factor returns from original paper data Excel."""
    wb = openpyxl.load_workbook(DATA_DIR / 'VME_Original_Paper_Data.xlsx', data_only=True)
    ws = wb['VME Factors']
    header = [cell.value for cell in ws[15]]
    rows = []
    for r in range(16, ws.max_row + 1):
        row = [cell.value for cell in ws[r]]
        if row[0] is None:
            break
        rows.append(row)
    df = pd.DataFrame(rows, columns=header)
    df['DATE'] = pd.to_datetime(df['DATE'])
    df = df.set_index('DATE')
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    wb.close()
    print(f"  Factors: {len(df)} months  {df.index[0]:%Y-%m} ~ {df.index[-1]:%Y-%m}")
    print(f"  Columns: {list(df.columns)}")
    return df


# ============================================================
# 2. Macro regressors
# ============================================================
def load_macro_regressors(factors_index):
    """
    Build the 6 macro regressors used in Table III.

    Returns a DataFrame aligned to monthly frequency (month-end).
    Variables:
      cons_growth  — long-run consumption growth proxy (quarterly, forward-filled)
      recession    — NBER recession dummy (1=trough, 0=expansion)
      gdp_growth   — quarterly real per capita GDP growth, forward-filled to monthly
      market       — US stock market excess return (Mkt-RF from FF, decimal)
      term         — 10Y - 3M T-bill spread (monthly mean), proxy for TERM bond factor
      default      — BAA - AAA spread level, proxy for DEF bond factor
    """

    # ── NBER recession ──────────────────────────────────────────────
    df_rec = pd.read_csv(DATA_DIR / 'nber_recession.csv', parse_dates=['observation_date'])
    df_rec = df_rec.set_index('observation_date')
    df_rec.index = df_rec.index + pd.offsets.MonthEnd(0)
    recession = df_rec['USREC'].astype(float)

    # ── GDP growth (quarterly → monthly forward-fill) ────────────────
    df_gdp = pd.read_csv(DATA_DIR / 'gdp_real.csv', parse_dates=['observation_date'])
    df_gdp = df_gdp.set_index('observation_date')
    df_gdp['GDPC1'] = pd.to_numeric(df_gdp['GDPC1'], errors='coerce')
    # quarterly per-capita growth (log)
    gdp_q = np.log(df_gdp['GDPC1']).diff()
    gdp_q.index = gdp_q.index + pd.offsets.MonthEnd(0)
    # resample to month-end then forward-fill within the quarter
    gdp_monthly = gdp_q.resample('ME').last().ffill()

    # ── Consumption growth (proxy: same as GDP — quarterly log growth) ─
    # The paper uses 3-year future consumption growth (sum of 12 future quarters).
    # Since we cannot compute future values, we use contemporaneous quarterly
    # log consumption growth as a proxy, forward-filled to monthly.
    # (Same GDP series used as proxy — matches paper's Table III direction check)
    cons_growth = gdp_monthly.copy()   # proxy

    # ── Market return (FF Mkt-RF, percent → decimal) ────────────────
    ff_path = DATA_DIR / 'ff_factors_tmp' / 'F-F_Research_Data_Factors.csv'
    ff_raw = open(ff_path).read()
    # find the data section (skip text header)
    lines = ff_raw.splitlines()
    data_lines = []
    annual_section = False
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        # annual section starts after a blank-ish separator
        if 'Annual Factors' in stripped or (data_lines and len(stripped.split(',')) == 1):
            annual_section = True
        if annual_section:
            continue
        parts = stripped.split(',')
        if len(parts) == 5:
            try:
                yyyymm = int(parts[0].strip())
                if 190001 <= yyyymm <= 209912:
                    data_lines.append(parts)
            except ValueError:
                continue

    ff_df = pd.DataFrame(data_lines, columns=['yyyymm', 'Mkt-RF', 'SMB', 'HML', 'RF'])
    ff_df['yyyymm'] = ff_df['yyyymm'].astype(int)
    ff_df['date'] = pd.to_datetime(ff_df['yyyymm'].astype(str), format='%Y%m') + pd.offsets.MonthEnd(0)
    ff_df = ff_df.set_index('date')
    for c in ['Mkt-RF', 'SMB', 'HML', 'RF']:
        ff_df[c] = pd.to_numeric(ff_df[c], errors='coerce')
    market = ff_df['Mkt-RF'] / 100.0   # percent → decimal

    # ── TERM: 10Y-3M spread (daily → monthly mean) ─────────────────
    df_term = pd.read_csv(DATA_DIR / 'term_spread.csv', parse_dates=['observation_date'])
    df_term = df_term.set_index('observation_date')
    df_term['T10Y3M'] = pd.to_numeric(df_term['T10Y3M'], errors='coerce')
    term = df_term['T10Y3M'].resample('ME').mean()
    term.index = term.index + pd.offsets.MonthEnd(0)
    term = term / 100.0   # percent → decimal (consistent units with DEF)

    # ── DEF: BAA - AAA spread level ─────────────────────────────────
    df_baa = pd.read_csv(DATA_DIR / 'baa_yield.csv', parse_dates=['observation_date'])
    df_aaa = pd.read_csv(DATA_DIR / 'aaa_yield.csv', parse_dates=['observation_date'])
    df_baa = df_baa.set_index('observation_date')
    df_aaa = df_aaa.set_index('observation_date')
    df_baa.index = df_baa.index + pd.offsets.MonthEnd(0)
    df_aaa.index = df_aaa.index + pd.offsets.MonthEnd(0)
    df_baa['BAA'] = pd.to_numeric(df_baa['BAA'], errors='coerce')
    df_aaa['AAA'] = pd.to_numeric(df_aaa['AAA'], errors='coerce')
    default = (df_baa['BAA'] - df_aaa['AAA']) / 100.0   # percent → decimal

    # ── Assemble ────────────────────────────────────────────────────
    macro = pd.DataFrame({
        'cons_growth': cons_growth,
        'recession':   recession,
        'gdp_growth':  gdp_monthly,
        'market':      market,
        'term':        term,
        'default':     default,
    })

    # Align to factors sample, then forward-fill quarterly gaps
    macro = macro.reindex(factors_index)
    macro = macro.ffill()

    # NOTE: TERM (T10Y3M) is only available from 1982-01 onward.
    # The paper uses the Fama-French TERM bond factor return (long-term govt
    # bond return minus T-bill return), which covers the full 1972+ sample.
    # Using the yield spread as a proxy loses the 1972-1981 period.
    # We keep NaNs for TERM before 1982 — they will be handled by dropna per regression.
    print(f"\n  Macro regressors coverage (relative to factor sample start {factors_index[0]:%Y-%m}):")
    for c in macro.columns:
        nn = macro[c].notna().sum()
        first = macro[c].first_valid_index()
        print(f"    {c:15s}: {nn} non-NA obs  (from {first:%Y-%m} if not None)"
              if first is not None else f"    {c:15s}: 0 non-NA obs")

    return macro


# ============================================================
# 3. Table III
# ============================================================
def run_table_III(factors_df, macro):
    """
    Table III: Multivariate macro regressions.

    Single OLS per specification with ALL 6 variables simultaneously.
    HAC (Newey-West) standard errors with 6 lags.

    Specifications:
      US stocks  : VAL^US, MOM^US
      Global stk : VAL^SS, MOM^SS  (VAL^SS = stock-only VAL)
      Nonstock   : VAL^AA, MOM^AA  (AA = all assets minus stocks)
      All assets : VAL, MOM
    """

    specs = [
        ('U.S. Stocks',     'VAL^US', 'MOM^US'),
        ('Global Stocks',   'VAL^SS', 'MOM^SS'),
        ('Nonstock Assets', 'VAL^AA', 'MOM^AA'),
        ('All Assets',      'VAL',    'MOM'),
    ]

    var_labels = ['cons_growth', 'recession', 'gdp_growth', 'market', 'term', 'default']
    display_labels = {
        'cons_growth': 'Consumption growth',
        'recession':   'Recession',
        'gdp_growth':  'GDP growth',
        'market':      'Market',
        'term':        'TERM',
        'default':     'DEF',
    }

    print("\n" + "=" * 100)
    print("TABLE III: Multivariate Macro Regressions  (OLS, HAC 6 lags)")
    print("  r_t = α + β₁·ConsGrowth + β₂·Recession + β₃·GDP + β₄·Market + β₅·TERM + β₆·DEF + ε")
    print("=" * 100)

    # Paper target numbers for comparison
    paper_III = {
        'VAL^US': {
            'cons_growth': (0.0004, 2.06),
            'recession':   (-0.0068, -1.06),
            'gdp_growth':  (-0.0050, -1.75),
            'market':      (-0.3435, -1.46),
            'term':        (0.2038, 2.64),
            'default':     (0.2650, 5.25),
            'R2':          0.131,
        },
        'MOM^US': {
            'cons_growth': (0.0001, 0.33),
            'recession':   (-0.0056, -0.73),
            'gdp_growth':  (0.0019, 0.57),
            'market':      (-0.0709, 0.80),
            'term':        (0.0141, 0.27),
            'default':     (-0.3572, -8.37),
            'R2':          0.059,
        },
        'VAL': {
            'cons_growth': (0.0001, 1.01),
            'recession':   (0.0043, 1.35),
            'gdp_growth':  (-0.0006, -0.45),
            'market':      (-0.0068, -0.32),
            'term':        (0.0551, 2.22),
            'default':     (0.0824, 3.40),
            'R2':          0.029,
        },
        'MOM': {
            'cons_growth': (0.0001, 0.43),
            'recession':   (-0.0072, -2.26),
            'gdp_growth':  (0.0020, -1.29),
            'market':      (-0.0068, -0.23),
            'term':        (-0.0240, -0.56),
            'default':     (-0.1480, -2.84),
            'R2':          0.047,
        },
    }

    all_results = {}

    for spec_name, val_col, mom_col in specs:
        print(f"\n{'─'*100}")
        print(f"  {spec_name}  —  Value: {val_col}  |  Momentum: {mom_col}")
        print(f"{'─'*100}")
        print(f"  {'Variable':22s}  {'Val coef':>10s} {'Val t':>8s}  {'Mom coef':>10s} {'Mom t':>8s}", end='')
        if val_col in paper_III:
            print(f"  |  {'Paper Val':>10s} {'Paper t':>8s}  {'Paper Mom':>10s} {'Paper t':>8s}", end='')
        print()
        print(f"  {'─'*100}")

        results = {}
        for role, col in [('val', val_col), ('mom', mom_col)]:
            if col not in factors_df.columns:
                print(f"  [WARN] {col} not found in factors")
                results[role] = None
                continue

            y_raw = factors_df[col].dropna()
            # Intersect with macro
            common = y_raw.index.intersection(macro.dropna(how='any').index)
            if len(common) < 60:
                print(f"  [WARN] {col}: only {len(common)} overlapping obs")
                results[role] = None
                continue

            # Use only rows where ALL regressors are available
            macro_sub = macro.loc[common, var_labels].dropna(how='any')
            common2 = y_raw.index.intersection(macro_sub.index)
            if len(common2) < 60:
                print(f"  [WARN] {col}: only {len(common2)} complete obs after dropna")
                results[role] = None
                continue
            y = y_raw[common2]
            X = sm.add_constant(macro_sub.loc[common2])
            reg = sm.OLS(y, X).fit(cov_type='HAC', cov_kwds={'maxlags': 6})
            results[role] = reg

        if results.get('val') is None and results.get('mom') is None:
            continue

        # Print row by row
        val_reg = results.get('val')
        mom_reg = results.get('mom')

        for v in var_labels:
            lbl = display_labels[v]
            vc = val_reg.params.get(v, np.nan) if val_reg else np.nan
            vt = val_reg.tvalues.get(v, np.nan) if val_reg else np.nan
            mc = mom_reg.params.get(v, np.nan) if mom_reg else np.nan
            mt = mom_reg.tvalues.get(v, np.nan) if mom_reg else np.nan

            row = f"  {lbl:22s}  {vc:10.4f} {vt:8.2f}  {mc:10.4f} {mt:8.2f}"
            if val_col in paper_III:
                pvc, pvt = paper_III[val_col].get(v, (np.nan, np.nan))
                pmc, pmt = paper_III[mom_col].get(v, (np.nan, np.nan)) if mom_col in paper_III else (np.nan, np.nan)
                row += f"  |  {pvc:10.4f} ({pvt:6.2f})  {pmc:10.4f} ({pmt:6.2f})"
            print(row)

        # R²
        vr2 = val_reg.rsquared if val_reg else np.nan
        mr2 = mom_reg.rsquared if mom_reg else np.nan
        row_r2 = f"  {'R²':22s}  {vr2:10.3f} {'':8s}  {mr2:10.3f} {'':8s}"
        if val_col in paper_III:
            pvr2 = paper_III[val_col].get('R2', np.nan)
            pmr2 = paper_III[mom_col].get('R2', np.nan) if mom_col in paper_III else np.nan
            row_r2 += f"  |  {pvr2:10.3f} {'':>8s}  {pmr2:10.3f}"
        print(row_r2)

        N_val = len(val_reg.fittedvalues) if val_reg else 0
        N_mom = len(mom_reg.fittedvalues) if mom_reg else 0
        print(f"  N (val={N_val}, mom={N_mom})")

        all_results[spec_name] = results

    return all_results


# ============================================================
# 4. Table IV helpers
# ============================================================
def compute_ar2_residuals(series: pd.Series, name: str) -> pd.Series:
    """
    Fit AR(2) on series, return residuals as liquidity shocks.
    series must be a clean (dropna) pd.Series with DatetimeIndex.
    """
    y = series.dropna()
    if len(y) < 30:
        print(f"  [WARN] AR(2) for {name}: only {len(y)} obs")
        return pd.Series(dtype=float)

    # AR(2): y_t = c + phi1*y_{t-1} + phi2*y_{t-2} + eps_t
    X = pd.DataFrame({
        'const': 1.0,
        'lag1':  y.shift(1),
        'lag2':  y.shift(2),
    }, index=y.index).dropna()
    y_aligned = y[X.index]
    ols = sm.OLS(y_aligned, X).fit()
    resid = ols.resid
    resid.name = f'{name}_shock'
    print(f"  AR(2) {name}: N={len(y_aligned)}, R²={ols.rsquared:.3f}, "
          f"phi1={ols.params['lag1']:.3f}, phi2={ols.params['lag2']:.3f}")
    return resid


def load_liquidity_shocks(macro):
    """
    Build liquidity shock measures for Table IV.

    1. TED spread AR(2) residual  → funding liquidity shock
    2. Pastor-Stambaugh innovation (column 3 of PS file) → market liquidity
    3. 'All PC' shock (first principal component of TED + PS) as proxy for paper's PC
    """
    shocks = {}

    # ── TED spread AR(2) ────────────────────────────────────────────
    df_ted = pd.read_csv(DATA_DIR / 'ted_spread.csv', parse_dates=['observation_date'])
    df_ted = df_ted.set_index('observation_date')
    df_ted['TEDRATE'] = pd.to_numeric(df_ted['TEDRATE'], errors='coerce')
    ted_monthly = df_ted['TEDRATE'].resample('ME').last()
    ted_monthly.index = ted_monthly.index + pd.offsets.MonthEnd(0)
    ted_clean = ted_monthly.dropna()
    ted_shock = compute_ar2_residuals(ted_clean, 'TED')
    shocks['TED spread'] = ted_shock

    # ── Pastor-Stambaugh: use column 3 (innovation) directly ────────
    ps_rows = []
    with open(DATA_DIR / 'pastor_stambaugh_liquidity.txt') as f:
        for line in f:
            line = line.strip()
            if line.startswith('%') or not line:
                continue
            parts = line.split()
            if len(parts) < 3:
                continue
            try:
                yyyymm = int(float(parts[0]))
                year, month = yyyymm // 100, yyyymm % 100
                if 1960 <= year <= 2025 and 1 <= month <= 12:
                    innov = float(parts[2])
                    traded = float(parts[3]) if len(parts) >= 4 else np.nan
                    ps_rows.append({
                        'date': pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0),
                        'innovation': innov if innov != -99 else np.nan,
                        'traded': traded if traded != -99 else np.nan,
                    })
            except (ValueError, IndexError):
                continue

    ps_df = pd.DataFrame(ps_rows).set_index('date')
    # PS column 3 is already the AR(8) innovation (equation 8) — use directly
    shocks['Pastor-Stambaugh'] = ps_df['innovation'].dropna()
    print(f"  PS Liquidity: {len(shocks['Pastor-Stambaugh'])} obs "
          f"{shocks['Pastor-Stambaugh'].index[0]:%Y-%m} ~ {shocks['Pastor-Stambaugh'].index[-1]:%Y-%m}")

    # ── All PC: first principal component of TED shock + PS innovation ─
    # Align to common dates
    shock_panel = pd.DataFrame({
        'ted_shock': shocks['TED spread'],
        'ps_innov':  shocks['Pastor-Stambaugh'],
    }).dropna()

    if len(shock_panel) >= 30:
        from sklearn.decomposition import PCA
        scaler_mean = shock_panel.mean()
        scaler_std  = shock_panel.std()
        normed = (shock_panel - scaler_mean) / scaler_std
        pca = PCA(n_components=1)
        pc1 = pca.fit_transform(normed)[:, 0]
        pc1_series = pd.Series(pc1, index=shock_panel.index, name='All_PC_shock')
        # Flip sign so that positive = more liquidity (matches paper convention)
        # Paper: positive liquidity shock → positive momentum beta
        if pc1_series.corr(shocks['Pastor-Stambaugh'][shock_panel.index]) < 0:
            pc1_series = -pc1_series
        shocks['All PC'] = pc1_series
        print(f"  All PC (PCA of TED+PS): {len(pc1_series)} obs, "
              f"explained variance: {pca.explained_variance_ratio_[0]:.1%}")
    else:
        print("  [WARN] Not enough data for All PC")

    return shocks


def run_table_IV(factors_df, macro, liq_shocks):
    """
    Table IV: Liquidity risk exposures.

    For each liquidity shock:
      r_t = α + β·liq_shock_t + (Table III macro controls) + ε_t

    Dependent variables per Panel A:
      Value (VAL^US), Momentum (MOM^US), Combo (50/50), Val-Mom (VAL^US - MOM^US)

    Uses HAC standard errors (6 lags).
    """

    print("\n" + "=" * 100)
    print("TABLE IV: Liquidity Risk Exposures  (OLS, HAC 6 lags)")
    print("  r_t = α + β·Liq_shock + β_macro·Controls + ε")
    print("=" * 100)

    # Paper targets (Panel A)
    paper_IV = {
        'TED spread': {
            'val': (-0.0052, -1.44), 'mom': (0.0129, 3.07),
            'combo': (0.0061, 2.13), 'valmom': (-0.0180, -2.62),
        },
        'Pastor-Stambaugh': {
            'val': (0.0034, 0.32), 'mom': (0.0107, 0.89),
            'combo': (0.0159, 1.93), 'valmom': (-0.0074, -0.37),
        },
        'All PC': {
            'val': (-0.0154, -2.84), 'mom': (0.0195, 2.96),
            'combo': (0.0043, 1.09), 'valmom': (-0.0349, -3.17),
        },
    }

    macro_controls = ['recession', 'gdp_growth', 'market', 'term', 'default']

    # Build dependent variable series
    val_us  = factors_df.get('VAL^US', factors_df.get('VAL', None))
    mom_us  = factors_df.get('MOM^US', factors_df.get('MOM', None))
    if val_us is None or mom_us is None:
        print("  [ERROR] VAL^US or MOM^US not found")
        return

    val_us  = val_us.dropna()
    mom_us  = mom_us.dropna()
    common_ym = val_us.index.intersection(mom_us.index)
    val_us  = val_us[common_ym]
    mom_us  = mom_us[common_ym]
    combo   = 0.5 * val_us + 0.5 * mom_us
    valmom  = val_us - mom_us   # Val minus Mom spread

    dep_vars = {
        'val':    ('Value (VAL^US)',     val_us),
        'mom':    ('Momentum (MOM^US)',  mom_us),
        'combo':  ('Combo (50/50)',       combo),
        'valmom': ('Val-Mom spread',     valmom),
    }

    print(f"\n  {'Liquidity measure':22s}  "
          f"{'Val β':>8s} {'Val t':>8s}  "
          f"{'Mom β':>8s} {'Mom t':>8s}  "
          f"{'Combo β':>8s} {'Combo t':>8s}  "
          f"{'V-M β':>8s} {'V-M t':>8s}")
    print(f"  {'─'*100}")

    for shock_name, shock_series in liq_shocks.items():
        shock_clean = shock_series.dropna()

        # Build regressors: liq_shock + macro controls
        reg_df = pd.DataFrame({'liq_shock': shock_clean})
        for ctl in macro_controls:
            if ctl in macro.columns:
                reg_df[ctl] = macro[ctl]
        reg_df = reg_df.dropna()

        row_coefs = []
        row_tstats = []

        for role, (dep_label, dep_series) in dep_vars.items():
            common = dep_series.dropna().index.intersection(reg_df.dropna(how='any').index)
            if len(common) < 30:
                row_coefs.append(np.nan)
                row_tstats.append(np.nan)
                continue

            y = dep_series[common]
            X = sm.add_constant(reg_df.loc[common].dropna(how='any'))
            y = y[X.index]   # re-align after any remaining NaN drop
            reg = sm.OLS(y, X).fit(cov_type='HAC', cov_kwds={'maxlags': 6})
            beta = reg.params.get('liq_shock', np.nan)
            t    = reg.tvalues.get('liq_shock', np.nan)
            row_coefs.append(beta)
            row_tstats.append(t)

        # Format output line
        vc, vt, mc, mt, cc, ct, dc, dt = (*row_coefs, *row_tstats) if len(row_coefs) == 4 else [np.nan]*8
        # unpack properly
        vc, mc, cc, dc = row_coefs
        vt, mt, ct, dt = row_tstats

        line = (f"  {shock_name:22s}  "
                f"{vc:8.4f} {vt:8.2f}  "
                f"{mc:8.4f} {mt:8.2f}  "
                f"{cc:8.4f} {ct:8.2f}  "
                f"{dc:8.4f} {dt:8.2f}")
        print(line)

        # Paper comparison
        if shock_name in paper_IV:
            p = paper_IV[shock_name]
            pvc, pvt = p['val']
            pmc, pmt = p['mom']
            pcc, pct = p['combo']
            pdc, pdt = p['valmom']
            paper_line = (f"  {'  → Paper':22s}  "
                          f"{pvc:8.4f} ({pvt:5.2f})  "
                          f"{pmc:8.4f} ({pmt:5.2f})  "
                          f"{pcc:8.4f} ({pct:5.2f})  "
                          f"{pdc:8.4f} ({pdt:5.2f})")
            print(paper_line)
            print()

    # Also run Panel B style: using global VAL/MOM
    print(f"\n  Panel B style — Global (VAL, MOM):")
    print(f"  {'─'*100}")
    print(f"  {'Liquidity measure':22s}  "
          f"{'Val β':>8s} {'Val t':>8s}  "
          f"{'Mom β':>8s} {'Mom t':>8s}  "
          f"{'Combo β':>8s} {'Combo t':>8s}  "
          f"{'V-M β':>8s} {'V-M t':>8s}")
    print(f"  {'─'*100}")

    gval = factors_df['VAL'].dropna()
    gmom = factors_df['MOM'].dropna()
    gcommon = gval.index.intersection(gmom.index)
    gval = gval[gcommon]; gmom = gmom[gcommon]
    gcombo = 0.5 * gval + 0.5 * gmom
    gvalmom = gval - gmom

    dep_global = {
        'val':    ('Global Value',  gval),
        'mom':    ('Global Mom',    gmom),
        'combo':  ('Global Combo',  gcombo),
        'valmom': ('Global V-M',    gvalmom),
    }

    for shock_name, shock_series in liq_shocks.items():
        shock_clean = shock_series.dropna()
        reg_df = pd.DataFrame({'liq_shock': shock_clean})
        for ctl in macro_controls:
            if ctl in macro.columns:
                reg_df[ctl] = macro[ctl]
        reg_df = reg_df.dropna()

        row_coefs = []
        row_tstats = []
        for role, (dep_label, dep_series) in dep_global.items():
            common = dep_series.dropna().index.intersection(reg_df.dropna(how='any').index)
            if len(common) < 30:
                row_coefs.append(np.nan); row_tstats.append(np.nan)
                continue
            y = dep_series[common]
            X = sm.add_constant(reg_df.loc[common].dropna(how='any'))
            y = y[X.index]
            reg = sm.OLS(y, X).fit(cov_type='HAC', cov_kwds={'maxlags': 6})
            row_coefs.append(reg.params.get('liq_shock', np.nan))
            row_tstats.append(reg.tvalues.get('liq_shock', np.nan))

        vc, mc, cc, dc = row_coefs
        vt, mt, ct, dt = row_tstats
        print(f"  {shock_name:22s}  "
              f"{vc:8.4f} {vt:8.2f}  "
              f"{mc:8.4f} {mt:8.2f}  "
              f"{cc:8.4f} {ct:8.2f}  "
              f"{dc:8.4f} {dt:8.2f}")


# ============================================================
# 5. Main
# ============================================================
def main():
    print("=" * 100)
    print("  VME Replication — Table III & Table IV (v2)")
    print("  Asness, Moskowitz, Pedersen (2013) Journal of Finance")
    print("=" * 100)

    print("\n[1] Loading factor data...")
    factors_df = load_factors()

    print("\n[2] Building macro regressors...")
    macro = load_macro_regressors(factors_df.index)

    print("\n[3] Running Table III: Multivariate Macro Regressions...")
    run_table_III(factors_df, macro)

    print("\n[4] Building liquidity shocks for Table IV...")
    try:
        liq_shocks = load_liquidity_shocks(macro)
    except ImportError:
        print("  [WARN] sklearn not available; skipping All PC")
        # Rebuild without PCA
        liq_shocks = {}
        df_ted = pd.read_csv(DATA_DIR / 'ted_spread.csv', parse_dates=['observation_date'])
        df_ted = df_ted.set_index('observation_date')
        df_ted['TEDRATE'] = pd.to_numeric(df_ted['TEDRATE'], errors='coerce')
        ted_monthly = df_ted['TEDRATE'].resample('ME').last()
        ted_monthly.index = ted_monthly.index + pd.offsets.MonthEnd(0)
        liq_shocks['TED spread'] = compute_ar2_residuals(ted_monthly.dropna(), 'TED')
        ps_rows = []
        with open(DATA_DIR / 'pastor_stambaugh_liquidity.txt') as f:
            for line in f:
                line = line.strip()
                if line.startswith('%') or not line:
                    continue
                parts = line.split()
                if len(parts) >= 3:
                    try:
                        yyyymm = int(float(parts[0]))
                        year, month = yyyymm // 100, yyyymm % 100
                        if 1960 <= year <= 2025 and 1 <= month <= 12:
                            innov = float(parts[2])
                            ps_rows.append({'date': pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0),
                                            'innovation': innov if innov != -99 else np.nan})
                    except (ValueError, IndexError):
                        continue
        ps_df = pd.DataFrame(ps_rows).set_index('date')
        liq_shocks['Pastor-Stambaugh'] = ps_df['innovation'].dropna()

    print("\n[5] Running Table IV: Liquidity Risk Exposures...")
    run_table_IV(factors_df, macro, liq_shocks)

    print("\n" + "=" * 100)
    print("  Done.")
    print("=" * 100)


if __name__ == '__main__':
    main()
