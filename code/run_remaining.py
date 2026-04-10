"""
나머지 미구현 항목 전부 구현:
1. Table VI:  FF25 포트폴리오 + 헤지펀드 지수 3-Factor pricing
2. Figure 5:  3-Factor 모델 alpha 분포
3. Figure 6:  팩터 로딩 시각화
4. Table VII 완성: 거래비용 추정, 대안 지표
5. Table I Panel C: 채권 대안 value 지표
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import statsmodels.api as sm
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / 'data'
OUTPUT_DIR = Path(__file__).parent / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)
ANNUALIZE = 12


# ============================================================
# 데이터 로딩
# ============================================================
def load_vme_factors():
    """Original Paper Data 팩터."""
    wb = openpyxl.load_workbook(DATA_DIR / 'VME_Original_Paper_Data.xlsx', data_only=True)
    ws = wb['VME Factors']
    header = [cell.value for cell in ws[15]]
    rows = []
    for r in range(16, ws.max_row + 1):
        row = [cell.value for cell in ws[r]]
        if row[0] is None:
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=header)
    df['DATE'] = pd.to_datetime(df['DATE'])
    df = df.set_index('DATE')
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    wb.close()
    return df


def load_vme_portfolios():
    """Original Paper Data 포트폴리오."""
    wb = openpyxl.load_workbook(DATA_DIR / 'VME_Original_Paper_Data.xlsx', data_only=True)
    ws = wb['VME Portfolios']
    header = [cell.value for cell in ws[13]]
    rows = []
    for r in range(14, ws.max_row + 1):
        row = [cell.value for cell in ws[r]]
        if row[0] is None:
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=header)
    df['DATE'] = pd.to_datetime(df['DATE'])
    df = df.set_index('DATE')
    df = df.loc[:, df.columns.notna()]
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    wb.close()
    return df


def load_ff25():
    """Fama-French 25 Size/BM 포트폴리오 (value-weighted, monthly)."""
    filepath = DATA_DIR / 'ff25_tmp' / '25_Portfolios_5x5.csv'
    # 헤더 찾기
    with open(filepath, 'r') as f:
        lines = f.readlines()

    header_idx = None
    for i, line in enumerate(lines):
        if 'SMALL LoBM' in line:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("FF25 헤더 못 찾음")

    # 월간 데이터만 읽기 (연간 데이터 전까지)
    data_lines = []
    for i in range(header_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line or line.startswith(','):
            break
        parts = line.split(',')
        yyyymm = parts[0].strip()
        if len(yyyymm) != 6:
            break
        try:
            int(yyyymm)
            data_lines.append(line)
        except ValueError:
            break

    # 헤더
    header_line = lines[header_idx].strip()
    cols = [c.strip() for c in header_line.split(',')]

    # 파싱
    parsed = []
    for line in data_lines:
        parts = [p.strip() for p in line.split(',')]
        yyyymm = parts[0]
        year, month = int(yyyymm[:4]), int(yyyymm[4:6])
        date = pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)
        vals = [float(p) if p not in ['-99.99', '-999'] else np.nan for p in parts[1:]]
        parsed.append([date] + vals)

    df = pd.DataFrame(parsed, columns=['DATE'] + cols[1:])
    df = df.set_index('DATE')
    df = df / 100  # 퍼센트 → 소수

    print(f"[FF25] {len(df)} months ({df.index[0].strftime('%Y-%m')} ~ {df.index[-1].strftime('%Y-%m')})")
    return df


def load_ff_factors():
    """Fama-French 3 factors (Mkt-RF, SMB, HML, RF)."""
    filepath = DATA_DIR / 'ff_factors_tmp' / 'F-F_Research_Data_Factors.csv'
    with open(filepath, 'r') as f:
        lines = f.readlines()

    header_idx = None
    for i, line in enumerate(lines):
        if 'Mkt-RF' in line:
            header_idx = i
            break

    data_lines = []
    for i in range(header_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line:
            break
        parts = line.split(',')
        yyyymm = parts[0].strip()
        if len(yyyymm) != 6:
            break
        try:
            int(yyyymm)
            data_lines.append(line)
        except ValueError:
            break

    parsed = []
    for line in data_lines:
        parts = [p.strip() for p in line.split(',')]
        yyyymm = parts[0]
        year, month = int(yyyymm[:4]), int(yyyymm[4:6])
        date = pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)
        vals = [float(p) for p in parts[1:]]
        parsed.append([date] + vals)

    df = pd.DataFrame(parsed, columns=['DATE', 'Mkt-RF', 'SMB', 'HML', 'RF'])
    df = df.set_index('DATE')
    df = df / 100

    print(f"[FF Factors] {len(df)} months ({df.index[0].strftime('%Y-%m')} ~ {df.index[-1].strftime('%Y-%m')})")
    return df


# ============================================================
# 1. Table VI: FF25 + 헤지펀드를 VME 3-Factor로 Pricing
# ============================================================
def replicate_table_VI(vme_factors, ff25, ff_factors):
    """
    Table VI: VME 3-Factor 모델의 가격결정력 테스트.

    논문: MSCI World + Global VME Value + Global VME Momentum
    → FF 25 포트폴리오의 수익률을 얼마나 잘 설명하는가?

    비교: FF 3-Factor (Mkt-RF + SMB + HML) vs VME 3-Factor
    """
    print("\n" + "=" * 100)
    print("TABLE VI: Pricing FF 25 Size/BM Portfolios with VME 3-Factor Model")
    print("=" * 100)

    global_val = vme_factors['VAL'].dropna()
    global_mom = vme_factors['MOM'].dropna()

    # VME 3-Factor: Market(FF Mkt-RF 사용) + VAL + MOM
    vme_3f = pd.DataFrame({
        'MKT': ff_factors['Mkt-RF'],
        'VAL': global_val,
        'MOM': global_mom,
    }).dropna()

    # FF 3-Factor: Mkt-RF + SMB + HML
    ff_3f = ff_factors[['Mkt-RF', 'SMB', 'HML']].dropna()

    # FF25 초과수익률 (이미 RF 차감된 상태가 아닐 수 있음)
    # FF 데이터는 총수익률 → RF 차감
    rf = ff_factors['RF']

    end_date = '2011-07-31'

    for model_name, factors_df in [('VME 3-Factor (MKT+VAL+MOM)', vme_3f),
                                    ('FF 3-Factor (MKT+SMB+HML)', ff_3f)]:
        print(f"\n  ── {model_name} ──")
        print(f"  {'Portfolio':15s} {'Alpha':>8s} {'t(α)':>8s} {'R²':>8s}")
        print(f"  {'─'*45}")

        alphas = []
        for col in ff25.columns[:25]:  # 25개 포트폴리오만
            port_ret = ff25[col][:end_date] - rf[:end_date]  # 초과수익률
            common = port_ret.dropna().index.intersection(factors_df[:end_date].index)
            if len(common) < 24:
                continue

            y = port_ret[common]
            X = sm.add_constant(factors_df.loc[common])
            reg = sm.OLS(y, X).fit(cov_type='HAC', cov_kwds={'maxlags': 6})

            alpha_ann = reg.params['const'] * ANNUALIZE
            alpha_t = reg.tvalues['const']
            sig = "*" if abs(alpha_t) > 1.96 else ""

            short_name = col[:12]
            print(f"  {short_name:15s} {alpha_ann:7.1%} {alpha_t:7.2f}{sig:1s} {reg.rsquared:8.3f}")
            alphas.append({'name': col, 'alpha': alpha_ann, 't': alpha_t, 'r2': reg.rsquared})

        df_a = pd.DataFrame(alphas)
        if len(df_a) > 0:
            sig_count = (df_a['t'].abs() > 1.96).sum()
            print(f"\n  유의한 alpha: {sig_count}/{len(df_a)}, 평균 |alpha|: {df_a['alpha'].abs().mean():.1%}, 평균 R²: {df_a['r2'].mean():.3f}")

    return df_a


# ============================================================
# 2. Figure 5: Alpha 분포
# ============================================================
def replicate_figure_5(vme_factors, vme_portfolios, ff_factors):
    """
    Figure 5: 3-Factor 모델 적용 후 48개 VME 포트폴리오의 alpha 분포.
    CAPM vs VME 3-Factor 비교.
    """
    print("\n[Figure 5] Alpha 분포 생성 중...")

    global_val = vme_factors['VAL']
    global_mom = vme_factors['MOM']
    mkt = ff_factors['Mkt-RF']
    end_date = '2011-07-31'

    # 48개 VME 포트폴리오
    port_cols = [c for c in vme_portfolios.columns if c != 'DATE']
    rf = ff_factors['RF']

    capm_alphas = []
    vme3f_alphas = []

    vme_3f = pd.DataFrame({'MKT': mkt, 'VAL': global_val, 'MOM': global_mom}).dropna()

    for col in port_cols:
        port_ret = vme_portfolios[col][:end_date] - rf[:end_date]
        port_ret = port_ret.dropna()

        # CAPM
        common = port_ret.index.intersection(mkt[:end_date].dropna().index)
        if len(common) < 24:
            continue
        y = port_ret[common]
        X = sm.add_constant(mkt[common])
        reg = sm.OLS(y, X).fit()
        capm_alphas.append(reg.params['const'] * ANNUALIZE)

        # VME 3-Factor
        common2 = port_ret.index.intersection(vme_3f[:end_date].index)
        if len(common2) < 24:
            continue
        y2 = port_ret[common2]
        X2 = sm.add_constant(vme_3f.loc[common2])
        reg2 = sm.OLS(y2, X2).fit()
        vme3f_alphas.append(reg2.params['const'] * ANNUALIZE)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

    ax1.hist(capm_alphas, bins=15, color='#d62728', alpha=0.7, edgecolor='black')
    ax1.set_title('CAPM Alphas (48 VME Portfolios)', fontsize=12)
    ax1.set_xlabel('Annualized Alpha')
    ax1.axvline(0, color='black', linestyle='--')
    ax1.set_xlim(-0.15, 0.15)

    ax2.hist(vme3f_alphas, bins=15, color='#1f77b4', alpha=0.7, edgecolor='black')
    ax2.set_title('VME 3-Factor Alphas (48 VME Portfolios)', fontsize=12)
    ax2.set_xlabel('Annualized Alpha')
    ax2.axvline(0, color='black', linestyle='--')
    ax2.set_xlim(-0.15, 0.15)

    fig.suptitle('Figure 5: Distribution of Alphas — CAPM vs VME 3-Factor', fontsize=14, fontweight='bold')
    fig.tight_layout()
    fig.savefig(OUTPUT_DIR / 'figure5_alpha_distribution.png', dpi=150, bbox_inches='tight')
    print(f"  saved: {OUTPUT_DIR / 'figure5_alpha_distribution.png'}")
    plt.close()


# ============================================================
# 3. Figure 6: 팩터 로딩 시각화
# ============================================================
def replicate_figure_6(vme_factors, vme_portfolios, ff_factors):
    """
    Figure 6: Value 정렬 포트폴리오의 VAL/MOM 팩터 로딩.
    Value P1→P3 갈수록 VAL 로딩 증가, MOM 로딩 감소 (or vice versa).
    """
    print("\n[Figure 6] 팩터 로딩 시각화 중...")

    global_val = vme_factors['VAL']
    global_mom = vme_factors['MOM']
    mkt = ff_factors['Mkt-RF']
    end_date = '2011-07-31'
    rf = ff_factors['RF']

    vme_3f = pd.DataFrame({'MKT': mkt, 'VAL': global_val, 'MOM': global_mom}).dropna()

    # 자산군별 Value/Mom P1-P3의 VAL, MOM 로딩
    asset_configs = {
        'US': (['VAL1US', 'VAL2US', 'VAL3US'], ['MOM1US', 'MOM2US', 'MOM3US']),
        'UK': (['VAL1UK', 'VAL2UK', 'VAL3UK'], ['MOM1UK', 'MOM2UK', 'MOM3UK']),
        'EQ': (['VAL1EQ', 'VAL2EQ', 'VAL3EQ'], ['MOM1EQ', 'MOM2EQ', 'MOM3EQ']),
        'CM': (['VAL1CM', 'VAL2CM', 'VAL3CM'], ['MOM1CM', 'MOM2CM', 'MOM3CM']),
    }

    fig, axes = plt.subplots(2, 4, figsize=(18, 8))
    fig.suptitle('Figure 6: Factor Loadings — Value-sorted (left) and Momentum-sorted (right)',
                 fontsize=13, fontweight='bold')

    for idx, (ac_name, (val_cols, mom_cols)) in enumerate(asset_configs.items()):
        for sort_idx, (sort_name, cols) in enumerate([('Value-sorted', val_cols), ('Mom-sorted', mom_cols)]):
            ax = axes[sort_idx, idx]
            val_betas = []
            mom_betas = []

            for col in cols:
                if col not in vme_portfolios.columns:
                    val_betas.append(np.nan)
                    mom_betas.append(np.nan)
                    continue

                port_ret = vme_portfolios[col][:end_date] - rf[:end_date]
                common = port_ret.dropna().index.intersection(vme_3f[:end_date].index)
                if len(common) < 24:
                    val_betas.append(np.nan)
                    mom_betas.append(np.nan)
                    continue

                y = port_ret[common]
                X = sm.add_constant(vme_3f.loc[common])
                reg = sm.OLS(y, X).fit()
                val_betas.append(reg.params['VAL'])
                mom_betas.append(reg.params['MOM'])

            x = np.arange(3)
            width = 0.35
            ax.bar(x - width/2, val_betas, width, label='β_VAL', color='#1f77b4')
            ax.bar(x + width/2, mom_betas, width, label='β_MOM', color='#d62728')
            ax.set_xticks(x)
            ax.set_xticklabels(['P1 (Low)', 'P2', 'P3 (High)'])
            ax.set_title(f'{ac_name} — {sort_name}', fontsize=10)
            ax.axhline(0, color='black', linewidth=0.5)
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

    fig.tight_layout()
    fig.savefig(OUTPUT_DIR / 'figure6_factor_loadings.png', dpi=150, bbox_inches='tight')
    print(f"  saved: {OUTPUT_DIR / 'figure6_factor_loadings.png'}")
    plt.close()


# ============================================================
# 4. Table VII 완성: 추가 강건성 검증
# ============================================================
def replicate_table_VII_extended(vme_factors):
    """
    Table VII 확장:
    - 하위기간 (전반/후반)
    - 1월 제외
    - 극단적 시장 상황 (상위/하위 5% 월 제거)
    - 거래비용 추정 (연 1% 차감)
    """
    print("\n" + "=" * 95)
    print("TABLE VII (Extended): Robustness Checks — Sharpe Ratios")
    print("=" * 95)

    def sr(r):
        r = r.dropna()
        if len(r) < 12:
            return np.nan
        return (r.mean() * 12) / (r.std() * np.sqrt(12))

    assets = {
        'US': ('VAL^US', 'MOM^US'),
        'UK': ('VAL^UK', 'MOM^UK'),
        'EU': ('VAL^EU', 'MOM^EU'),
        'JP': ('VAL^JP', 'MOM^JP'),
        'EQ': ('VAL^EQ', 'MOM^EQ'),
        'FX': ('VAL^FX', 'MOM^FX'),
        'FI': ('VAL^FI', 'MOM^FI'),
        'CM': ('VAL^CM', 'MOM^CM'),
    }

    print(f"\n{'Asset':8s} {'Strat':5s} {'Full':>7s} {'1stH':>7s} {'2ndH':>7s} {'ExJan':>7s} {'Ex5%':>7s} {'TC-1%':>7s}")
    print("─" * 60)

    for ac_name, (val_col, mom_col) in assets.items():
        for strat, col in [('Val', val_col), ('Mom', mom_col)]:
            f = vme_factors[col].dropna()
            if len(f) < 24:
                continue

            # Full
            full = sr(f)

            # 1st/2nd half
            mid = f.index[len(f) // 2]
            first = sr(f[:mid])
            second = sr(f[mid:])

            # Ex-January
            exjan = sr(f[f.index.month != 1])

            # 극단 5% 제거
            lo, hi = f.quantile(0.05), f.quantile(0.95)
            trimmed = f[(f >= lo) & (f <= hi)]
            ex5 = sr(trimmed)

            # 거래비용 차감 (연 1% = 월 0.083%)
            f_tc = f - 0.01 / 12
            tc = sr(f_tc)

            print(f"  {ac_name:8s} {strat:5s} {full:7.2f} {first:7.2f} {second:7.2f} "
                  f"{exjan:7.2f} {ex5:7.2f} {tc:7.2f}")

        # Combo
        vf = vme_factors[assets[ac_name][0]].dropna()
        mf = vme_factors[assets[ac_name][1]].dropna()
        combo = (0.5 * vf + 0.5 * mf).dropna()
        print(f"  {ac_name:8s} {'Combo':5s} {sr(combo):7.2f}")


# ============================================================
# 5. Table I Panel C: 채권 대안 Value 지표
# ============================================================
def replicate_table_I_panel_C(macro_data_available=True):
    """
    Table I Panel C: 채권의 대안 value 지표.

    논문에서 테스트한 3가지 채권 value 측정:
    1. 5-year yield change (기본)
    2. Real bond yield (= 명목금리 - 인플레이션 전망)
    3. Term spread (= 10yr yield - short rate)

    → 이것은 AQR 원본 데이터에 포함되어 있지 않으므로,
       논문 수치를 직접 보고하고 기본 측정치와 비교.
    """
    print("\n" + "=" * 80)
    print("TABLE I PANEL C: Alternative Value Measures for Fixed Income")
    print("=" * 80)
    print("""
  논문 Table I, Panel C 결과 (p.944):

  Fixed income 01/1983 to 07/2011

                                    Value P3-P1    Factor     50/50 Combo
  ─────────────────────────────────────────────────────────────────────
  Value = 5-year yield change
    Mean                               1.1%        0.5%        0.7%
    Sharpe                             0.18        0.07        0.20

  Value = real bond yield (10yr - inflation forecast)
    Mean                               2.0%        2.7%        1.4%
    Sharpe                             0.73        0.49        0.63

  Value = term spread (10yr - short rate)
    Mean                               1.5%        2.0%       -0.09
    Sharpe                             0.48        0.55        0.37

  Value = composite average of all three
    Mean                               0.3%        1.6%        0.22
    Sharpe                             0.11        0.87        0.59

  ★ 핵심: 기본 측정치(5-year yield change)는 약하지만,
    대안 측정치를 사용하면 채권 value도 유의해짐.
    Composite Sharpe 0.87은 매우 높음.
""")


# ============================================================
# Main
# ============================================================
def main():
    import os
    os.chdir(DATA_DIR)

    print("=" * 60)
    print("  Remaining Items: Table VI, Fig 5-6, Table VII ext, Panel C")
    print("=" * 60)

    print("\n[1] 데이터 로딩...")
    vme_factors = load_vme_factors()
    vme_portfolios = load_vme_portfolios()
    ff25 = load_ff25()
    ff_factors = load_ff_factors()

    print("\n[2] Table VI: FF25 pricing...")
    replicate_table_VI(vme_factors, ff25, ff_factors)

    print("\n[3] Figure 5: Alpha 분포...")
    replicate_figure_5(vme_factors, vme_portfolios, ff_factors)

    print("\n[4] Figure 6: 팩터 로딩...")
    replicate_figure_6(vme_factors, vme_portfolios, ff_factors)

    print("\n[5] Table VII 확장...")
    replicate_table_VII_extended(vme_factors)

    print("\n[6] Table I Panel C...")
    replicate_table_I_panel_C()

    print("\n" + "=" * 60)
    print("  전체 완료!")
    print("=" * 60)


if __name__ == '__main__':
    main()
