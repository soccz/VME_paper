"""
Tables V, VI, VII 재현: Value and Momentum Everywhere (Asness, Moskowitz, Pedersen 2013)

Table V  (p.970): Fama-MacBeth Cross-Sectional Regressions
Table VI (p.972): Time-Series Asset Pricing Tests (GRS F-stat, Avg |α|, R², XS-R², Eig%)
Table VII(p.980): Dynamics — Sharpe ratios and correlations across economic environments
"""

import numpy as np
import pandas as pd
import statsmodels.api as sm
import scipy.stats as stats
import openpyxl
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')

DATA_DIR  = Path(__file__).parent.parent / 'data'
OUTPUT_DIR = Path(__file__).parent / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)

ANNUALIZE = 12
PAPER_END = '2011-07-31'
HAC_LAGS  = 6

# ──────────────────────────────────────────────────────────────────────────────
# 1. DATA LOADING
# ──────────────────────────────────────────────────────────────────────────────

def load_original_paper_factors():
    """
    VME_Original_Paper_Data.xlsx — VME Factors sheet.
    Header at row 15, data from row 16.
    Also reads STD of passive market from row 13.
    Returns (factors_df, passive_mkt_std_series).
    """
    wb = openpyxl.load_workbook(DATA_DIR / 'VME_Original_Paper_Data.xlsx', data_only=True)
    ws = wb['VME Factors']

    # Row 13: STD of passive market (label in col 0, values in cols 1..)
    std_row = [cell.value for cell in ws[13]]
    header15 = [cell.value for cell in ws[15]]

    # Align STD values with header columns
    passive_std = {}
    for i, colname in enumerate(header15):
        if colname is not None and colname != 'DATE' and i < len(std_row):
            val = std_row[i]
            if val is not None:
                try:
                    passive_std[colname] = float(val)
                except (TypeError, ValueError):
                    pass

    # Main data
    rows = []
    for r in range(16, ws.max_row + 1):
        row = [cell.value for cell in ws[r]]
        if row[0] is None:
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=header15)
    df['DATE'] = pd.to_datetime(df['DATE'])
    df = df.set_index('DATE')
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    wb.close()
    print(f"[Factors] {len(df)} months ({df.index[0]:%Y-%m} ~ {df.index[-1]:%Y-%m})")
    print(f"  Columns ({len(df.columns)}): {list(df.columns)}")
    print(f"  Passive MKT std entries: {len(passive_std)}")

    return df, passive_std


def load_original_paper_portfolios():
    """
    VME_Original_Paper_Data.xlsx — VME Portfolios sheet.
    Header at row 13, data from row 14.
    Returns DataFrame with 48 portfolio columns.
    """
    wb = openpyxl.load_workbook(DATA_DIR / 'VME_Original_Paper_Data.xlsx', data_only=True)
    ws = wb['VME Portfolios']
    header = [cell.value for cell in ws[13]]
    rows = []
    for r in range(14, ws.max_row + 1):
        row = [cell.value for cell in ws[r]]
        if row[0] is None:
            continue
        rows.append(row)
    # Only keep columns with actual headers (first 49 cols, rest are None)
    valid_cols = [(i, h) for i, h in enumerate(header) if h is not None]
    valid_indices = [i for i, _ in valid_cols]
    valid_headers = [h for _, h in valid_cols]

    # Re-extract rows with only valid columns
    rows_trimmed = []
    for row in rows:
        rows_trimmed.append([row[i] if i < len(row) else None for i in valid_indices])

    df = pd.DataFrame(rows_trimmed, columns=valid_headers)
    date_col = df.columns[0]
    df[date_col] = pd.to_datetime(df[date_col])
    df = df.set_index(date_col)
    df.index.name = 'DATE'
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    df = df.dropna(how='all')
    wb.close()
    print(f"[Portfolios] {len(df)} months ({df.index[0]:%Y-%m} ~ {df.index[-1]:%Y-%m})")
    print(f"  Columns ({len(df.columns)}): {list(df.columns)}")
    return df


def load_ff_factors():
    """FF factors: Mkt-RF, SMB, HML, RF (in decimal, dates as month-end)."""
    filepath = DATA_DIR / 'ff_factors_tmp' / 'F-F_Research_Data_Factors.csv'
    with open(filepath, 'r') as f:
        lines = f.readlines()
    header_idx = None
    for i, line in enumerate(lines):
        if 'Mkt-RF' in line:
            header_idx = i
            break
    data_lines = []
    for i in range(header_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line:
            break
        parts = line.split(',')
        yyyymm = parts[0].strip()
        if len(yyyymm) != 6:
            break
        try:
            int(yyyymm)
            data_lines.append(line)
        except ValueError:
            break
    parsed = []
    for line in data_lines:
        parts = [p.strip() for p in line.split(',')]
        yyyymm = parts[0]
        year, month = int(yyyymm[:4]), int(yyyymm[4:6])
        date = pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)
        vals = [float(p) for p in parts[1:]]
        parsed.append([date] + vals)
    df = pd.DataFrame(parsed, columns=['DATE', 'Mkt-RF', 'SMB', 'HML', 'RF'])
    df = df.set_index('DATE')
    df = df / 100  # percent to decimal
    print(f"[FF] {len(df)} months ({df.index[0]:%Y-%m} ~ {df.index[-1]:%Y-%m})")
    return df


def load_ff25():
    """FF 25 Size/BM portfolios (in %, value-weighted)."""
    # find header row containing 'SMALL LoBM'
    raw = open(DATA_DIR / 'ff25_tmp' / '25_Portfolios_5x5.csv').read()
    lines = raw.split('\n')
    skip = 0
    for i, line in enumerate(lines):
        if 'SMALL LoBM' in line or 'Lo BM' in line or 'Small Lo BM' in line:
            skip = i
            break

    ff25 = pd.read_csv(DATA_DIR / 'ff25_tmp' / '25_Portfolios_5x5.csv',
                       skiprows=skip, header=0)
    ff25.columns = [c.strip() for c in ff25.columns]
    date_col = ff25.columns[0]
    ff25 = ff25.rename(columns={date_col: 'DATE'})
    ff25 = ff25[ff25['DATE'].astype(str).str.match(r'^\d{6}$')]
    ff25['DATE'] = (pd.to_datetime(ff25['DATE'].astype(str), format='%Y%m')
                    + pd.offsets.MonthEnd(0))
    ff25 = ff25.set_index('DATE')
    # keep only first 25 columns (value-weighted returns section)
    ff25 = ff25.iloc[:, :25]
    for col in ff25.columns:
        ff25[col] = pd.to_numeric(ff25[col], errors='coerce') / 100.0
    # Replace -99.99 / -999 with NaN
    ff25 = ff25.replace(-99.99, np.nan).replace(-999, np.nan)
    ff25 = ff25.dropna(how='all')
    print(f"[FF25] {len(ff25)} months ({ff25.index[0]:%Y-%m} ~ {ff25.index[-1]:%Y-%m})")
    return ff25


def load_macro_factors():
    """
    Load TERM and DEF spreads as monthly factors.
    TERM = 10Y - 3M Treasury spread (T10Y3M, daily → monthly last)
    DEF  = BAA - AAA yield spread
    """
    # TERM spread (daily → monthly last)
    term = pd.read_csv(DATA_DIR / 'term_spread.csv', parse_dates=['observation_date'])
    term = term.set_index('observation_date')
    term['T10Y3M'] = pd.to_numeric(term['T10Y3M'], errors='coerce')
    term_m = term['T10Y3M'].resample('ME').last()
    term_m.index = term_m.index + pd.offsets.MonthEnd(0)

    # DEF spread (monthly, already month-begin → month-end)
    baa = pd.read_csv(DATA_DIR / 'baa_yield.csv', parse_dates=['observation_date'])
    baa = baa.set_index('observation_date')
    baa.index = baa.index + pd.offsets.MonthEnd(0)
    baa['BAA'] = pd.to_numeric(baa['BAA'], errors='coerce')

    aaa = pd.read_csv(DATA_DIR / 'aaa_yield.csv', parse_dates=['observation_date'])
    aaa = aaa.set_index('observation_date')
    aaa.index = aaa.index + pd.offsets.MonthEnd(0)
    aaa['AAA'] = pd.to_numeric(aaa['AAA'], errors='coerce')

    def_spread = baa['BAA'] - aaa['AAA']

    return term_m.rename('TERM'), def_spread.rename('DEF')


def load_ps_liquidity():
    """Pastor-Stambaugh (2003) traded liquidity factor."""
    ps_lines = []
    with open(DATA_DIR / 'pastor_stambaugh_liquidity.txt', 'r') as f:
        for line in f:
            line = line.strip()
            if line.startswith('%') or not line:
                continue
            parts = line.split()
            if len(parts) >= 4:
                try:
                    yyyymm = int(float(parts[0]))
                    year   = yyyymm // 100
                    month  = yyyymm % 100
                    if 1960 <= year <= 2025 and 1 <= month <= 12:
                        traded_val = float(parts[3])
                        ps_lines.append({
                            'date': pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0),
                            'ps_liq': traded_val if traded_val != -99 else np.nan,
                        })
                except (ValueError, IndexError):
                    continue
    ps_df = pd.DataFrame(ps_lines).set_index('date')
    print(f"[PS Liq] {len(ps_df)} months ({ps_df.index[0]:%Y-%m} ~ {ps_df.index[-1]:%Y-%m})")
    return ps_df['ps_liq']


# ──────────────────────────────────────────────────────────────────────────────
# 2. HELPER FUNCTIONS
# ──────────────────────────────────────────────────────────────────────────────

def newey_west_tstat(y, x_with_const, maxlags=HAC_LAGS):
    """OLS with Newey-West HAC standard errors. Returns (params, tstats, rsq)."""
    reg = sm.OLS(y, x_with_const).fit(cov_type='HAC', cov_kwds={'maxlags': maxlags})
    return reg.params.values, reg.tvalues.values, reg.rsquared


def rolling_beta(y, x, window=60):
    """
    Rolling univariate OLS beta of y on x (with constant).
    Returns Series of betas aligned to y.index (NaN for first window-1 obs).
    """
    betas = np.full(len(y), np.nan)
    xa = np.array(x)
    ya = np.array(y)
    for t in range(window - 1, len(y)):
        yi = ya[t - window + 1: t + 1]
        xi = xa[t - window + 1: t + 1]
        mask = ~(np.isnan(yi) | np.isnan(xi))
        if mask.sum() < 30:
            continue
        yi_c = yi[mask]
        xi_c = xi[mask]
        X = np.column_stack([np.ones(len(xi_c)), xi_c])
        try:
            coef, _, _, _ = np.linalg.lstsq(X, yi_c, rcond=None)
            betas[t] = coef[1]
        except np.linalg.LinAlgError:
            pass
    return pd.Series(betas, index=y.index)


def fama_macbeth(port_rets, factor_betas_dict, dates):
    """
    Fama-MacBeth (1973) cross-sectional regression.

    Parameters
    ----------
    port_rets   : DataFrame (T × N) — portfolio excess returns
    factor_betas_dict : OrderedDict {name: DataFrame(T × N)} — rolling betas
    dates       : pd.DatetimeIndex — common date range

    Returns
    -------
    DataFrame of {coeff, tstat} for each factor + intercept
    """
    factor_names = list(factor_betas_dict.keys())
    N = port_rets.shape[1]
    gamma_series = {f: [] for f in ['intercept'] + factor_names}
    valid_dates  = []

    for t in dates:
        if t not in port_rets.index:
            continue
        r_t = port_rets.loc[t].values  # (N,)

        # Build cross-sectional X: N × (1 + K)
        X_cols = [np.ones(N)]
        ok = np.ones(N, dtype=bool)
        for fname in factor_names:
            bdf = factor_betas_dict[fname]
            if t not in bdf.index:
                ok[:] = False
                break
            b = bdf.loc[t].values
            X_cols.append(b)
            ok &= ~np.isnan(b)

        ok &= ~np.isnan(r_t)
        if ok.sum() < len(factor_names) + 2:
            continue

        X_t = np.column_stack(X_cols)[ok]
        y_t = r_t[ok]
        try:
            coef, _, _, _ = np.linalg.lstsq(X_t, y_t, rcond=None)
        except np.linalg.LinAlgError:
            continue

        valid_dates.append(t)
        gamma_series['intercept'].append(coef[0])
        for k, fname in enumerate(factor_names):
            gamma_series[fname].append(coef[k + 1])

    if not valid_dates:
        return None

    results = {}
    for k, v in gamma_series.items():
        arr = np.array(v)
        mean_g = arr.mean()
        # Newey-West t-stat on gamma time series
        T = len(arr)
        se = newey_west_se(arr, lags=HAC_LAGS)
        t_stat = mean_g / se if se > 0 else np.nan
        results[k] = {'coeff': mean_g, 'tstat': t_stat, 'n_months': T}

    return results


def newey_west_se(series, lags=HAC_LAGS):
    """Newey-West standard error of the mean of a 1-D series."""
    T = len(series)
    mu = series.mean()
    e  = series - mu
    var = np.dot(e, e) / T
    for j in range(1, lags + 1):
        w = 1 - j / (lags + 1)
        cov_j = np.dot(e[j:], e[:-j]) / T
        var += 2 * w * cov_j
    return np.sqrt(max(var, 0.0) / T)


def grs_test(alphas, residuals, factor_rets):
    """
    Gibbons, Ross, Shanken (1989) test.

    Parameters
    ----------
    alphas      : (N,) array
    residuals   : (T × N) array
    factor_rets : (T × K) array (demeaned optional — we demean internally)

    Returns
    -------
    (F_stat, p_value)
    """
    T, N = residuals.shape
    K    = factor_rets.shape[1]

    # Residual covariance (T-K-1 denominator)
    Sigma = (residuals.T @ residuals) / (T - K - 1)

    # Factor mean and covariance
    mu_f = factor_rets.mean(axis=0)               # (K,)
    Omega = np.cov(factor_rets.T, ddof=1)          # (K × K)
    if K == 1:
        Omega = Omega.reshape(1, 1)

    # GRS statistic
    try:
        Sigma_inv = np.linalg.inv(Sigma)
        Omega_inv = np.linalg.inv(Omega)
        quad_a = float(alphas @ Sigma_inv @ alphas)
        quad_f = float(mu_f @ Omega_inv @ mu_f)
        F = (T / N) * ((T - N - K) / (T - K - 1)) * quad_a / (1 + quad_f)
        p_val = 1 - stats.f.cdf(F, N, T - N - K)
    except np.linalg.LinAlgError:
        F, p_val = np.nan, np.nan

    return F, p_val


def ts_asset_pricing(port_rets, factors, model_name):
    """
    Time-series asset pricing for a given model.

    Parameters
    ----------
    port_rets : DataFrame (T × N) — portfolio returns (already excess)
    factors   : DataFrame (T × K) — factor returns

    Returns
    -------
    dict: grs_f, grs_p, avg_abs_alpha, avg_ts_r2, xs_r2, eig_pct
    """
    common = port_rets.index.intersection(factors.index)
    R = port_rets.loc[common].values         # (T × N)
    F = factors.loc[common].values           # (T × K)

    T, N = R.shape
    K    = F.shape[1]

    # Mask portfolios with too many NaNs
    valid_cols = np.where(np.sum(np.isnan(R), axis=0) < T * 0.3)[0]
    R = R[:, valid_cols]
    N = R.shape[1]

    alphas   = np.zeros(N)
    betas    = np.zeros((N, K))
    resids   = np.zeros((T, N))
    ts_r2    = np.zeros(N)
    pred_mu  = np.zeros(N)

    for i in range(N):
        yi = R[:, i]
        mask = ~np.isnan(yi)
        if mask.sum() < 30:
            alphas[i] = np.nan
            ts_r2[i] = np.nan
            continue

        Fi_rows = ~np.any(np.isnan(F[mask]), axis=1)
        y_i = yi[mask][Fi_rows]
        F_i = F[mask][Fi_rows]
        X_i = np.column_stack([np.ones(len(y_i)), F_i])

        coef, _, _, _ = np.linalg.lstsq(X_i, y_i, rcond=None)
        alphas[i] = coef[0]
        betas[i]  = coef[1:]

        y_hat = X_i @ coef
        ss_res = np.sum((y_i - y_hat) ** 2)
        ss_tot = np.sum((y_i - y_i.mean()) ** 2)
        ts_r2[i] = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan

        resid_full = np.full(T, np.nan)
        idx_positions = np.where(mask)[0][Fi_rows]
        resid_full[idx_positions] = y_i - y_hat
        resids[:, i] = resid_full

        # Predicted mean return
        pred_mu[i] = np.nanmean(F @ coef[1:])  # β̂ × factor_means

    # GRS test (drop NaN alphas/residuals)
    valid = ~np.isnan(alphas)
    if valid.sum() < 2:
        grs_f, grs_p = np.nan, np.nan
    else:
        alpha_v = alphas[valid]
        resid_v = resids[:, valid]
        F_nonan = F[~np.any(np.isnan(F), axis=1)]
        T_v     = (~np.any(np.isnan(resid_v), axis=1)).sum()
        # Use complete rows only
        complete_rows = ~np.any(np.isnan(resid_v), axis=1) & ~np.any(np.isnan(F), axis=1)
        if complete_rows.sum() < N + K + 5:
            grs_f, grs_p = np.nan, np.nan
        else:
            grs_f, grs_p = grs_test(alpha_v, resid_v[complete_rows], F[complete_rows])

    avg_abs_alpha = np.nanmean(np.abs(alphas))   # monthly
    avg_ts_r2     = np.nanmean(ts_r2)

    # Cross-sectional R²: regress actual mean returns on predicted mean returns
    actual_mu = np.nanmean(R, axis=0)
    valid2 = ~np.isnan(actual_mu) & ~np.isnan(pred_mu)
    if valid2.sum() > 2:
        slope, intercept, r, p, se = stats.linregress(pred_mu[valid2], actual_mu[valid2])
        xs_r2 = r ** 2
    else:
        xs_r2 = np.nan

    # Eig%: fraction of total variance explained by model-implied covariance
    # model-implied: B Ω_F B' + Σ_e  but paper uses just B Ω_F B'
    try:
        mu_f  = np.nanmean(F, axis=0)
        Omega_F = np.cov(F[~np.any(np.isnan(F), axis=1)].T, ddof=1)
        if K == 1:
            Omega_F = Omega_F.reshape(1, 1)
        B = betas[valid]
        cov_model = B @ Omega_F @ B.T    # N × N (signal part)

        # Sample covariance of R
        R_valid = R[:, valid]
        complete_r = ~np.any(np.isnan(R_valid), axis=1)
        cov_sample = np.cov(R_valid[complete_r].T, ddof=1)

        eig_model  = np.sum(np.linalg.eigvalsh(cov_model))
        eig_sample = np.sum(np.linalg.eigvalsh(cov_sample))
        eig_pct = eig_model / eig_sample if eig_sample > 0 else np.nan
    except Exception:
        eig_pct = np.nan

    return {
        'model':           model_name,
        'grs_f':           grs_f,
        'grs_p':           grs_p,
        'avg_abs_alpha':   avg_abs_alpha,   # monthly
        'avg_ts_r2':       avg_ts_r2,
        'xs_r2':           xs_r2,
        'eig_pct':         eig_pct,
        'n_portfolios':    N,
        'n_months':        T,
    }


def sharpe_ratio(ret_series, annualize=True):
    """Annualised Sharpe ratio from monthly returns."""
    r = ret_series.dropna()
    if len(r) < 12:
        return np.nan
    mu  = r.mean()
    sig = r.std(ddof=1)
    if sig == 0:
        return np.nan
    sr = mu / sig
    return sr * np.sqrt(12) if annualize else sr


# ──────────────────────────────────────────────────────────────────────────────
# 3. TABLE V: Fama-MacBeth Cross-Sectional Regressions
# ──────────────────────────────────────────────────────────────────────────────

def replicate_table_V(factors_df, portfolios_df, ff_df):
    """
    Table V (p.970): Fama-MacBeth regressions.

    Cross-sectional regression each month:
        R_i,t = γ₀ + γ₁β_liq,i + γ₂β_MKT,i + γ₃β_TERM,i + γ₄β_DEF,i
                   + γ₅β_VAL,i + γ₆β_MOM,i + ε_i,t

    Betas estimated via rolling 60-month univariate regressions.
    γ are reported as time-series means with NW t-stats (6 lags).
    """
    print("\n" + "=" * 100)
    print("TABLE V: Fama-MacBeth Cross-Sectional Regressions")
    print("  Cross-section each month: R_i = γ₀ + Σ γ_k β_{k,i} + ε_i")
    print("  Betas: rolling 60-month univariate OLS  |  γ t-stats: Newey-West 6 lags")
    print("=" * 100)

    # ----- Restrict to paper period -----
    end = pd.Timestamp(PAPER_END)

    # Get factor columns for all asset classes (individual VAL/MOM factors)
    # Paper uses 48 portfolios. We'll use the 48 portfolio returns from portfolios_df.
    port_cols = [c for c in portfolios_df.columns if c is not None and str(c).strip()]
    ports = portfolios_df[port_cols].copy().loc[:end]

    # Load macro factors
    term_m, def_m = load_macro_factors()
    ps_liq = load_ps_liquidity()

    # Market: Mkt-RF from FF (global proxy)
    mkt = ff_df['Mkt-RF'].copy().loc[:end]

    # VAL and MOM everywhere (from original paper)
    val_ew = factors_df['VAL'].copy().loc[:end]
    mom_ew = factors_df['MOM'].copy().loc[:end]

    # Build common date range (after rolling window burns 60 months)
    all_dates = ports.index.intersection(mkt.index)
    all_dates = all_dates[all_dates <= end]

    # Available factors
    avail_factors = {}

    # PS Liquidity (traded) — may be limited
    liq_common = ps_liq.dropna()
    if len(liq_common) > 60:
        avail_factors['β_liq']  = liq_common
    avail_factors['β_mkt']  = mkt
    avail_factors['β_term'] = term_m.reindex(mkt.index).dropna()
    avail_factors['β_def']  = def_m.reindex(mkt.index).dropna()
    avail_factors['β_val']  = val_ew
    avail_factors['β_mom']  = mom_ew

    print(f"\n  Available factors: {list(avail_factors.keys())}")
    print(f"  Portfolios: {len(port_cols)}  |  Date range: {all_dates[0]:%Y-%m} ~ {all_dates[-1]:%Y-%m}")

    # ----- Compute rolling 60-month betas for each portfolio × each factor -----
    print("  Computing rolling betas (window=60)...")
    WINDOW = 60
    beta_dfs = {}  # {factor_name: DataFrame(T × N)}

    N_ports = len(port_cols)
    for fname, fseries in avail_factors.items():
        beta_mat = pd.DataFrame(index=all_dates, columns=port_cols, dtype=float)
        for pcol in port_cols:
            port_s = ports[pcol].dropna()
            # Align
            common_idx = port_s.index.intersection(fseries.dropna().index)
            if len(common_idx) < WINDOW:
                continue
            y = port_s.reindex(all_dates)
            x = fseries.reindex(all_dates)
            betas = rolling_beta(y, x, window=WINDOW)
            beta_mat[pcol] = betas
        beta_dfs[fname] = beta_mat
        print(f"    {fname}: done ({int(beta_mat.notna().any(axis=1).sum())} months with data)")

    # Port returns aligned to all_dates
    R_ports = ports.reindex(all_dates)

    # ----- Run Fama-MacBeth for different model specifications -----
    specs = [
        ("Row 1: β_liq only",             ['β_liq']),
        ("Row 2: β_mkt only",              ['β_mkt']),
        ("Row 3: β_liq + β_mkt",           ['β_liq', 'β_mkt']),
        ("Row 4: β_liq + macro + β_mkt",   ['β_liq', 'β_mkt', 'β_term', 'β_def']),
        ("Row 5: full model (all)",        ['β_liq', 'β_mkt', 'β_term', 'β_def', 'β_val', 'β_mom']),
        ("Row 5b: β_mkt + β_val + β_mom",  ['β_mkt', 'β_val', 'β_mom']),
        ("Row 6: β_val + β_mom only",      ['β_val', 'β_mom']),
    ]

    fm_results = []
    for spec_name, factor_list in specs:
        # Skip specs with unavailable factors
        missing = [f for f in factor_list if f not in beta_dfs]
        if missing:
            print(f"\n  SKIP {spec_name}: missing {missing}")
            continue

        selected_betas = {f: beta_dfs[f] for f in factor_list}
        res = fama_macbeth(R_ports, selected_betas, all_dates)
        if res is None:
            print(f"\n  {spec_name}: no valid dates")
            continue

        fm_results.append((spec_name, factor_list, res))

        print(f"\n  {spec_name}  (n_months={res['intercept']['n_months']})")
        print(f"  {'Factor':15s} {'γ (monthly)':>14s} {'t-stat':>10s}")
        print(f"  {'─'*45}")
        print(f"  {'intercept':15s} {res['intercept']['coeff']:14.5f} {res['intercept']['tstat']:10.2f}")
        for fname in factor_list:
            if fname in res:
                coeff = res[fname]['coeff']
                tstat = res[fname]['tstat']
                sig = '**' if abs(tstat) > 2.57 else ('*' if abs(tstat) > 1.96 else '')
                print(f"  {fname:15s} {coeff:14.5f} {tstat:9.2f}{sig}")

    # Paper comparison
    print("\n" + "-" * 80)
    print("  PAPER COMPARISON (Table V, p.970):")
    print("  Row 1: β_liq only       → coeff=0.0024, t=3.05")
    print("  Row 4: +macro+mkt       → β_liq coeff=0.0005, t=0.56; β_mkt coeff=0.0029, t=2.58")
    print("  Row 5: full model       → β_liq=0.0016, t=1.38; β_val=0.0031, t=3.96; β_mom=0.003, t=3.53")
    print("  (Note: paper uses PS traded liquidity factor-mimicking portfolio, we use raw series)")
    print("-" * 80)

    return fm_results


# ──────────────────────────────────────────────────────────────────────────────
# 4. TABLE VI: Time-Series Asset Pricing Tests
# ──────────────────────────────────────────────────────────────────────────────

def replicate_table_VI(factors_df, portfolios_df, ff_df):
    """
    Table VI Panel A (p.972): Asset pricing tests on 48 VME portfolios.

    Models:
    1. CAPM (Global market = EW avg of all VME portfolios)
    2. Mkt + VAL + MOM (3-factor)
    3. Mkt + VAL only
    4. Mkt + MOM only
    5. FF 3-Factor (Mkt-RF, SMB, HML)
    6. FF 4-Factor (+ UMD from FF if available)
    7. Mkt + TERM + DEF (macro)

    Statistics: GRS F-stat, p-value, Avg|α| (monthly → shown as monthly),
                Avg TS-R², XS-R², Eig%
    """
    print("\n" + "=" * 100)
    print("TABLE VI Panel A: Time-Series Asset Pricing Tests (48 VME Portfolios)")
    print("  Statistics: GRS F / p-value / Avg|α| / Avg TS-R² / XS-R² / Eig%")
    print("=" * 100)

    end = pd.Timestamp(PAPER_END)

    # 48 portfolio returns
    port_cols = [c for c in portfolios_df.columns if c is not None and str(c).strip()]
    ports = portfolios_df[port_cols].copy().loc[:end]

    # VAL and MOM everywhere
    val_ew = factors_df['VAL'].loc[:end]
    mom_ew = factors_df['MOM'].loc[:end]

    # FF factors
    mkt_ff = ff_df['Mkt-RF'].loc[:end]
    smb    = ff_df['SMB'].loc[:end]
    hml    = ff_df['HML'].loc[:end]

    # Macro factors
    term_m, def_m = load_macro_factors()
    term_m = term_m.loc[:end]
    def_m  = def_m.loc[:end]

    # Global market proxy: EW average of all 48 portfolios
    ports_arr = ports.values.copy().astype(float)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        mkt_global = ports.mean(axis=1)   # simple EW average

    # Build model specifications
    models = []

    # 1. CAPM (global market)
    mkt_common = mkt_global.dropna()
    models.append(('CAPM (Global EW Mkt)',
                   pd.DataFrame({'Mkt': mkt_common})))

    # 2. Mkt + VAL + MOM
    combined = pd.concat([mkt_common, val_ew, mom_ew], axis=1).dropna()
    combined.columns = ['Mkt', 'VAL', 'MOM']
    models.append(('Mkt + VAL + MOM (3-factor)', combined))

    # 3. Mkt + VAL only
    mv = pd.concat([mkt_common, val_ew], axis=1).dropna()
    mv.columns = ['Mkt', 'VAL']
    models.append(('Mkt + VAL only', mv))

    # 4. Mkt + MOM only
    mm = pd.concat([mkt_common, mom_ew], axis=1).dropna()
    mm.columns = ['Mkt', 'MOM']
    models.append(('Mkt + MOM only', mm))

    # 5. FF CAPM (Mkt-RF only)
    ff_mkt = pd.DataFrame({'Mkt-RF': mkt_ff}).dropna()
    models.append(('FF CAPM', ff_mkt))

    # 6. FF 3-factor
    ff3 = pd.concat([mkt_ff, smb, hml], axis=1).dropna()
    ff3.columns = ['Mkt-RF', 'SMB', 'HML']
    models.append(('FF 3-Factor', ff3))

    # 7. Mkt + TERM + DEF (macro model)
    macro_f = pd.concat([mkt_common, term_m, def_m], axis=1).dropna()
    macro_f.columns = ['Mkt', 'TERM', 'DEF']
    models.append(('Mkt + TERM + DEF', macro_f))

    # ----- Run tests -----
    results = []
    print(f"\n  {'Model':<30s} {'GRS F':>8s} {'p-val':>8s} {'Avg|α|':>10s} {'Avg TS-R²':>10s} {'XS-R²':>8s} {'Eig%':>8s} {'N':>5s}")
    print(f"  {'─'*100}")

    for model_name, factor_df in models:
        res = ts_asset_pricing(ports, factor_df, model_name)
        results.append(res)
        avg_a  = res['avg_abs_alpha'] * 100  # monthly, in percent
        ts_r2  = res['avg_ts_r2']
        xs_r2  = res['xs_r2']
        eig    = res['eig_pct']
        grs_f  = res['grs_f']
        grs_p  = res['grs_p']
        N      = res['n_portfolios']
        T      = res['n_months']

        print(f"  {model_name:<30s} {grs_f:8.2f} {grs_p:8.3f} {avg_a:10.4f}% {ts_r2:10.3f} {xs_r2:8.3f} {eig:8.3f}  N={N} T={T}")

    # Paper comparison
    print("\n" + "-" * 100)
    print("  PAPER Table VI Panel A (p.972):")
    print(f"  {'Model':<30s} {'GRS F':>8s} {'p-val':>8s} {'Avg|α|':>10s} {'Avg TS-R²':>10s} {'XS-R²':>8s} {'Eig%':>8s}")
    print(f"  {'─'*85}")
    paper_rows = [
        ('CAPM',                          6.02, 0.000, 0.35, 0.40, 0.52, 0.57),
        ('Mkt+VAL+MOM',                   3.72, 0.000, 0.18, 0.68, 0.72, 0.84),
        ('Mkt+VAL only',                  3.80, 0.000, 0.28, 0.42, 0.43, 0.68),
        ('Mkt+MOM only',                  6.59, 0.000, 0.39, 0.30, 0.44, 0.47),
        ('FF 3-Factor',                   7.18, 0.000, 0.35, 0.31, 0.50, 0.53),
    ]
    for (name, gf, gp, aa, tsr, xsr, eig) in paper_rows:
        print(f"  {name:<30s} {gf:8.2f} {gp:8.3f} {aa:10.2f}% {tsr:10.2f} {xsr:8.2f} {eig:8.2f}")

    return results


# ──────────────────────────────────────────────────────────────────────────────
# 5. TABLE VII: Dynamics
# ──────────────────────────────────────────────────────────────────────────────

def replicate_table_VII(factors_df):
    """
    Table VII Panel A (p.980): Sharpe ratios and correlations across sub-periods.

    For each sub-period:
    - SR(Val): Sharpe ratio of global VAL factor
    - SR(Mom): Sharpe ratio of global MOM factor
    - SR(50/50): Sharpe of 0.5×VAL + 0.5×MOM
    - ρ(Val,Val): avg pairwise correlation among individual asset-class value strategies
    - ρ(Mom,Mom): avg pairwise correlation among individual asset-class mom strategies
    - ρ(Val,Mom): avg correlation between each val and each mom strategy (cross)
    """
    print("\n" + "=" * 100)
    print("TABLE VII Panel A: Sharpe Ratios and Correlations across Sub-Periods")
    print("=" * 100)

    end = pd.Timestamp(PAPER_END)

    # Global VAL and MOM
    val_ew = factors_df['VAL'].loc[:end].dropna()
    mom_ew = factors_df['MOM'].loc[:end].dropna()

    # Individual asset-class value and momentum factors
    # Identify columns: VAL^XX and MOM^XX patterns
    val_cols = [c for c in factors_df.columns
                if str(c).startswith('VAL^') and c != 'VAL']
    mom_cols = [c for c in factors_df.columns
                if str(c).startswith('MOM^') and c != 'MOM']

    print(f"  Global VAL/MOM: {val_ew.index[0]:%Y-%m} ~ {val_ew.index[-1]:%Y-%m}")
    print(f"  Individual VAL factors: {val_cols}")
    print(f"  Individual MOM factors: {mom_cols}")

    # Sub-periods (Table VII Panel A)
    sub_periods = [
        ('Full period (1972-2011)',  '1972-01-01', '2011-07-31'),
        ('1st half (1972-1991)',     '1972-01-01', '1991-12-31'),
        ('2nd half (1992-2010)',     '1992-01-01', '2010-12-31'),
        ('Pre-08/1998',             '1972-01-01', '1998-07-31'),
        ('Post-08/1998',            '1998-08-01', '2011-07-31'),
    ]

    print(f"\n  {'Period':<28s} {'SR Val':>8s} {'SR Mom':>8s} {'SR 50/50':>10s} {'ρ(V,V)':>8s} {'ρ(M,M)':>8s} {'ρ(V,M)':>8s}")
    print(f"  {'─'*85}")

    all_results = []

    for period_name, start_str, end_str in sub_periods:
        s = pd.Timestamp(start_str)
        e = pd.Timestamp(end_str)

        v_sub = val_ew.loc[s:e]
        m_sub = mom_ew.loc[s:e]

        sr_val  = sharpe_ratio(v_sub)
        sr_mom  = sharpe_ratio(m_sub)
        combo   = 0.5 * v_sub + 0.5 * m_sub
        sr_combo = sharpe_ratio(combo)

        # Individual factor correlations
        # Get individual VAL and MOM factor series for this period
        ind_val_series = []
        for vc in val_cols:
            s_f = factors_df[vc].loc[s:e].dropna()
            if len(s_f) > 24:
                ind_val_series.append(s_f)

        ind_mom_series = []
        for mc in mom_cols:
            s_f = factors_df[mc].loc[s:e].dropna()
            if len(s_f) > 24:
                ind_mom_series.append(s_f)

        # ρ(Val,Val): average pairwise correlation among val factors
        rho_vv = _avg_pairwise_corr(ind_val_series)
        # ρ(Mom,Mom): average pairwise correlation among mom factors
        rho_mm = _avg_pairwise_corr(ind_mom_series)
        # ρ(Val,Mom): average cross correlation
        rho_vm = _avg_cross_corr(ind_val_series, ind_mom_series)

        row = {
            'period': period_name,
            'sr_val': sr_val, 'sr_mom': sr_mom, 'sr_combo': sr_combo,
            'rho_vv': rho_vv, 'rho_mm': rho_mm, 'rho_vm': rho_vm,
            'n_val': len(ind_val_series), 'n_mom': len(ind_mom_series),
        }
        all_results.append(row)

        print(f"  {period_name:<28s} {sr_val:8.2f} {sr_mom:8.2f} {sr_combo:10.2f} "
              f"{rho_vv:8.2f} {rho_mm:8.2f} {rho_vm:8.2f}  "
              f"(nV={len(ind_val_series)}, nM={len(ind_mom_series)})")

    # Paper comparison
    print("\n" + "-" * 100)
    print("  PAPER Table VII Panel A (p.980):")
    print(f"  {'Period':<28s} {'SR Val':>8s} {'SR Mom':>8s} {'SR 50/50':>10s} {'ρ(V,V)':>8s} {'ρ(M,M)':>8s} {'ρ(V,M)':>8s}")
    print(f"  {'─'*85}")
    paper_vii = [
        ('1st half 1972-1991',   0.78, 0.90, 1.40, 0.31, 0.46, -0.44),
        ('2nd half 1992-2010',   0.68, 0.71, 1.43, 0.71, 0.77, -0.63),
        ('Pre-08/1998',          0.68, 1.02, 1.49, 0.16, 0.43, -0.51),
        ('Post-08/1998',         0.75, 0.72, 1.39, 0.64, 0.71, -0.55),
    ]
    for (name, sv, sm, sc, rvv, rmm, rvm) in paper_vii:
        print(f"  {name:<28s} {sv:8.2f} {sm:8.2f} {sc:10.2f} {rvv:8.2f} {rmm:8.2f} {rvm:8.2f}")

    return all_results


def _avg_pairwise_corr(series_list):
    """Average pairwise Pearson correlation among a list of pd.Series."""
    n = len(series_list)
    if n < 2:
        return np.nan
    corrs = []
    for i in range(n):
        for j in range(i + 1, n):
            common = series_list[i].index.intersection(series_list[j].index)
            if len(common) < 12:
                continue
            r = series_list[i][common].corr(series_list[j][common])
            if not np.isnan(r):
                corrs.append(r)
    return np.mean(corrs) if corrs else np.nan


def _avg_cross_corr(val_series_list, mom_series_list):
    """Average cross-correlation: each val × each mom."""
    corrs = []
    for vs in val_series_list:
        for ms in mom_series_list:
            common = vs.index.intersection(ms.index)
            if len(common) < 12:
                continue
            r = vs[common].corr(ms[common])
            if not np.isnan(r):
                corrs.append(r)
    return np.mean(corrs) if corrs else np.nan


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 100)
    print("Value and Momentum Everywhere — Tables V, VI, VII")
    print("=" * 100)

    # ----- Load data -----
    print("\n--- Loading data ---")
    factors_df, passive_std = load_original_paper_factors()
    portfolios_df           = load_original_paper_portfolios()
    ff_df                   = load_ff_factors()

    print(f"\nFactor columns: {list(factors_df.columns)}")
    print(f"Portfolio columns: {list(portfolios_df.columns)}")

    # ----- Table V -----
    print("\n\n" + "#" * 100)
    print("# TABLE V")
    print("#" * 100)
    try:
        tv_results = replicate_table_V(factors_df, portfolios_df, ff_df)
    except Exception as e:
        import traceback
        print(f"  ERROR in Table V: {e}")
        traceback.print_exc()

    # ----- Table VI -----
    print("\n\n" + "#" * 100)
    print("# TABLE VI")
    print("#" * 100)
    try:
        tvi_results = replicate_table_VI(factors_df, portfolios_df, ff_df)
    except Exception as e:
        import traceback
        print(f"  ERROR in Table VI: {e}")
        traceback.print_exc()

    # ----- Table VII -----
    print("\n\n" + "#" * 100)
    print("# TABLE VII")
    print("#" * 100)
    try:
        tvii_results = replicate_table_VII(factors_df)
    except Exception as e:
        import traceback
        print(f"  ERROR in Table VII: {e}")
        traceback.print_exc()

    print("\n" + "=" * 100)
    print("DONE")
    print("=" * 100)


if __name__ == '__main__':
    main()
