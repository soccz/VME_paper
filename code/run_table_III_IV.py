"""
Table III-IV 재현: 매크로/유동성 리스크와 Value/Momentum의 관계.

Table III: 경기변동, 유동성, 디폴트 리스크 변수와의 관계
Table IV: 유동성 리스크 분해 — Funding liquidity vs Market liquidity

논문 Section III (p.955~964):
- Value는 유동성 리스크에 음(-)으로 노출
- Momentum은 유동성 리스크에 양(+)으로 노출
- → Value와 Momentum의 음의 상관관계 일부 설명
"""

import numpy as np
import pandas as pd
import statsmodels.api as sm
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / 'data'
OUTPUT_DIR = Path(__file__).parent / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)
ANNUALIZE = 12


# ============================================================
# 1. 매크로 데이터 로딩
# ============================================================
def load_macro_data():
    """FRED + Pastor-Stambaugh 매크로 데이터 로딩."""
    macro = {}

    # --- NBER Recession (월간 더미) ---
    df = pd.read_csv(DATA_DIR / 'nber_recession.csv', parse_dates=['observation_date'])
    df = df.set_index('observation_date')
    df.index = df.index + pd.offsets.MonthEnd(0)
    macro['NBER Recession'] = df['USREC']
    print(f"  NBER Recession: {len(df)} months")

    # --- Default Spread (BAA - AAA) ---
    baa = pd.read_csv(DATA_DIR / 'baa_yield.csv', parse_dates=['observation_date']).set_index('observation_date')
    aaa = pd.read_csv(DATA_DIR / 'aaa_yield.csv', parse_dates=['observation_date']).set_index('observation_date')
    baa.index = baa.index + pd.offsets.MonthEnd(0)
    aaa.index = aaa.index + pd.offsets.MonthEnd(0)
    baa['BAA'] = pd.to_numeric(baa['BAA'], errors='coerce')
    aaa['AAA'] = pd.to_numeric(aaa['AAA'], errors='coerce')
    default_spread = baa['BAA'] - aaa['AAA']
    macro['Default Spread'] = default_spread
    # 변화분 (innovation)
    macro['Δ Default Spread'] = default_spread.diff()
    print(f"  Default Spread: {len(default_spread.dropna())} months")

    # --- TED Spread (일간 → 월말) ---
    ted = pd.read_csv(DATA_DIR / 'ted_spread.csv', parse_dates=['observation_date']).set_index('observation_date')
    ted['TEDRATE'] = pd.to_numeric(ted['TEDRATE'], errors='coerce')
    ted_monthly = ted['TEDRATE'].resample('ME').last()
    ted_monthly.index = ted_monthly.index + pd.offsets.MonthEnd(0)
    macro['TED Spread'] = ted_monthly
    macro['Δ TED Spread'] = ted_monthly.diff()
    print(f"  TED Spread: {len(ted_monthly.dropna())} months (from {ted_monthly.dropna().index[0].strftime('%Y-%m')})")

    # --- Term Spread (일간 → 월말) ---
    term = pd.read_csv(DATA_DIR / 'term_spread.csv', parse_dates=['observation_date']).set_index('observation_date')
    term['T10Y3M'] = pd.to_numeric(term['T10Y3M'], errors='coerce')
    term_monthly = term['T10Y3M'].resample('ME').last()
    term_monthly.index = term_monthly.index + pd.offsets.MonthEnd(0)
    macro['Term Spread'] = term_monthly
    print(f"  Term Spread: {len(term_monthly.dropna())} months")

    # --- GDP Growth (분기 → 월간 보간) ---
    gdp = pd.read_csv(DATA_DIR / 'gdp_real.csv', parse_dates=['observation_date']).set_index('observation_date')
    gdp['GDPC1'] = pd.to_numeric(gdp['GDPC1'], errors='coerce')
    gdp_growth = gdp['GDPC1'].pct_change()
    # 분기 → 월간 (전방 채움)
    gdp_monthly = gdp_growth.resample('ME').ffill()
    gdp_monthly.index = gdp_monthly.index + pd.offsets.MonthEnd(0)
    macro['GDP Growth'] = gdp_monthly
    print(f"  GDP Growth: {len(gdp_monthly.dropna())} months (quarterly interpolated)")

    # --- Consumption Growth (Nondurable + Services) ---
    try:
        nd = pd.read_csv(DATA_DIR / 'consumption_nondurable.csv', parse_dates=['observation_date']).set_index('observation_date')
        sv = pd.read_csv(DATA_DIR / 'consumption_services.csv', parse_dates=['observation_date']).set_index('observation_date')
        nd_col = [c for c in nd.columns if c != 'observation_date'][0]
        sv_col = [c for c in sv.columns if c != 'observation_date'][0]
        nd[nd_col] = pd.to_numeric(nd[nd_col], errors='coerce')
        sv[sv_col] = pd.to_numeric(sv[sv_col], errors='coerce')
        # 합산 후 성장률
        total = nd[nd_col].add(sv[sv_col], fill_value=0)
        total.index = total.index + pd.offsets.MonthEnd(0)
        macro['Consumption Growth'] = total.pct_change()
        print(f"  Consumption Growth: {len(total.dropna())} months")
    except Exception as e:
        print(f"  [WARN] Consumption Growth 로딩 실패: {e}")

    # --- Pastor-Stambaugh Liquidity Innovation ---
    try:
        ps_lines = []
        with open(DATA_DIR / 'pastor_stambaugh_liquidity.txt', 'r') as f:
            for line in f:
                line = line.strip()
                if line.startswith('%') or not line:
                    continue
                parts = line.split()
                if len(parts) >= 4:
                    try:
                        yyyymm = int(float(parts[0]))
                        year = yyyymm // 100
                        month = yyyymm % 100
                        if 1960 <= year <= 2025 and 1 <= month <= 12:
                            traded_val = float(parts[3])
                            ps_lines.append({
                                'date': pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0),
                                'level': float(parts[1]),
                                'innovation': float(parts[2]),
                                'traded': traded_val if traded_val != -99 else np.nan,
                            })
                    except (ValueError, IndexError):
                        continue

        ps_df = pd.DataFrame(ps_lines).set_index('date')
        macro['PS Liquidity Innovation'] = ps_df['innovation']
        macro['PS Traded Liquidity'] = ps_df['traded']
        print(f"  Pastor-Stambaugh: {len(ps_df)} months ({ps_df.index[0].strftime('%Y-%m')} ~ {ps_df.index[-1].strftime('%Y-%m')})")
    except Exception as e:
        print(f"  [WARN] Pastor-Stambaugh 로딩 실패: {e}")

    return macro


# ============================================================
# 2. 팩터 데이터 로딩 (Original Paper Data)
# ============================================================
def load_factors():
    """Original Paper Data에서 팩터 수익률 로딩."""
    wb = openpyxl.load_workbook(DATA_DIR / 'VME_Original_Paper_Data.xlsx', data_only=True)
    ws = wb['VME Factors']
    header = [cell.value for cell in ws[15]]
    rows = []
    for r in range(16, ws.max_row + 1):
        row = [cell.value for cell in ws[r]]
        if row[0] is None:
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=header)
    df['DATE'] = pd.to_datetime(df['DATE'])
    df = df.set_index('DATE')
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    wb.close()
    return df


# ============================================================
# 3. Table III: 매크로 변수 회귀
# ============================================================
def replicate_table_III(factors_df, macro):
    """
    Table III: 각 매크로 변수에 대한 Value/Momentum 노출도.

    논문 방법:
    - 단변량 시계열 회귀: r_t = a + b * macro_t + e_t
    - HAC 표준오차 (Newey-West, 6 lags)
    - 종속변수: Global Value, Global Momentum, Global Combo
    """
    print("\n" + "=" * 100)
    print("TABLE III: Macroeconomic and Liquidity Risk Exposures")
    print("  r_t = α + β × Macro_t + ε_t  (HAC t-stats, 6 lags)")
    print("=" * 100)

    # 글로벌 팩터
    global_val = factors_df['VAL'].dropna()
    global_mom = factors_df['MOM'].dropna()
    global_combo = (0.5 * global_val + 0.5 * global_mom).dropna()

    dep_vars = {
        'Global Value': global_val,
        'Global Momentum': global_mom,
        'Global Combo': global_combo,
    }

    # 개별 자산군 팩터도 추가
    asset_factors = {
        'US Val': factors_df.get('VAL^US'),
        'US Mom': factors_df.get('MOM^US'),
        'JP Val': factors_df.get('VAL^JP'),
        'JP Mom': factors_df.get('MOM^JP'),
    }

    macro_vars = [
        'NBER Recession', 'GDP Growth', 'Consumption Growth',
        'Default Spread', 'Δ Default Spread',
        'Term Spread', 'TED Spread', 'Δ TED Spread',
        'PS Liquidity Innovation', 'PS Traded Liquidity',
    ]

    results = []

    for macro_name in macro_vars:
        if macro_name not in macro:
            continue

        macro_series = macro[macro_name].dropna()
        if len(macro_series) < 24:
            continue

        print(f"\n  ── {macro_name} ──")
        print(f"  {'':25s} {'β':>10s} {'t(β)':>10s} {'R²':>8s}")
        print(f"  {'─'*60}")

        for dep_name, dep_series in dep_vars.items():
            if dep_series is None:
                continue
            dep = dep_series.dropna()
            common = dep.index.intersection(macro_series.index)
            if len(common) < 24:
                continue

            y = dep[common]
            X = sm.add_constant(macro_series[common])
            reg = sm.OLS(y, X).fit(cov_type='HAC', cov_kwds={'maxlags': 6})

            beta = reg.params.iloc[1]
            t_stat = reg.tvalues.iloc[1]
            r2 = reg.rsquared

            sig = "*" if abs(t_stat) > 1.96 else ""
            print(f"  {dep_name:25s} {beta:10.4f} {t_stat:9.2f}{sig:1s} {r2:8.3f}")

            results.append({
                'macro': macro_name, 'dep': dep_name,
                'beta': beta, 't_stat': t_stat, 'r2': r2,
            })

    return pd.DataFrame(results)


# ============================================================
# 4. Table IV: 유동성 리스크 분해
# ============================================================
def replicate_table_IV(factors_df, macro):
    """
    Table IV: 다변량 회귀 — 유동성 + 매크로 변수 동시 투입.

    논문의 핵심 발견:
    - Momentum은 유동성 리스크에 양(+)으로 노출
    - Value는 유동성 리스크에 음(-)으로 노출
    - → 둘의 음의 상관관계 일부 설명
    - 하지만 경제적 크기는 작음 (설명력 제한적)
    """
    print("\n" + "=" * 100)
    print("TABLE IV: Multivariate Regression — Liquidity Risk Decomposition")
    print("  r_t = α + β₁×Liquidity + β₂×Macro controls + ε_t")
    print("=" * 100)

    global_val = factors_df['VAL'].dropna()
    global_mom = factors_df['MOM'].dropna()
    global_combo = (0.5 * global_val + 0.5 * global_mom).dropna()

    # 회귀변수 구성
    regressors = {}
    for name in ['PS Liquidity Innovation', 'Δ TED Spread', 'Δ Default Spread',
                  'Term Spread', 'NBER Recession']:
        if name in macro:
            regressors[name] = macro[name].dropna()

    if len(regressors) < 2:
        print("  [WARN] 충분한 매크로 변수 없음")
        return

    # 공통 기간
    reg_df = pd.DataFrame(regressors)
    reg_df = reg_df.dropna()

    for dep_name, dep_series in [
        ('Global Value', global_val),
        ('Global Momentum', global_mom),
        ('Global Combo', global_combo),
    ]:
        common = dep_series.dropna().index.intersection(reg_df.index)
        if len(common) < 24:
            print(f"  {dep_name}: 관측수 부족 ({len(common)})")
            continue

        y = dep_series[common]
        X = sm.add_constant(reg_df.loc[common])

        reg = sm.OLS(y, X).fit(cov_type='HAC', cov_kwds={'maxlags': 6})

        print(f"\n  {dep_name} (N={len(common)}, R²={reg.rsquared:.3f})")
        print(f"  {'Variable':30s} {'Coef':>10s} {'t-stat':>10s}")
        print(f"  {'─'*55}")
        for var in reg.params.index:
            sig = "*" if abs(reg.tvalues[var]) > 1.96 else ""
            label = 'Intercept' if var == 'const' else var
            print(f"  {label:30s} {reg.params[var]:10.4f} {reg.tvalues[var]:9.2f}{sig:1s}")


# ============================================================
# 5. Figure 3-4: 매크로 리스크 시각화
# ============================================================
def replicate_figures_3_4(factors_df, macro):
    """
    Figure 3: 경기침체 시기의 Value/Momentum 성과
    Figure 4: 유동성 위기 시기 (1998 LTCM, 2008 GFC)
    """
    global_val = factors_df['VAL'].dropna()
    global_mom = factors_df['MOM'].dropna()
    global_combo = (0.5 * global_val + 0.5 * global_mom).dropna()

    fig, axes = plt.subplots(2, 1, figsize=(14, 10))

    # --- Figure 3: 누적수익률 + 경기침체 음영 ---
    ax = axes[0]
    ax.set_title('Figure 3: Cumulative Returns with NBER Recessions', fontsize=13, fontweight='bold')

    for series, label, color in [
        (global_val, 'Value', '#1f77b4'),
        (global_mom, 'Momentum', '#d62728'),
        (global_combo, 'Combo', '#2ca02c'),
    ]:
        cum = (1 + series).cumprod()
        ax.plot(cum.index, cum.values, label=label, color=color, linewidth=1.5)

    # 경기침체 음영
    if 'NBER Recession' in macro:
        rec = macro['NBER Recession']
        rec_periods = rec[rec == 1]
        if len(rec_periods) > 0:
            in_rec = False
            start = None
            for date, val in rec.items():
                if val == 1 and not in_rec:
                    start = date
                    in_rec = True
                elif val == 0 and in_rec:
                    ax.axvspan(start, date, alpha=0.2, color='gray')
                    in_rec = False
            if in_rec:
                ax.axvspan(start, rec.index[-1], alpha=0.2, color='gray')

    ax.set_yscale('log')
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    ax.set_ylabel('Cumulative Return')

    # --- Figure 4: TED Spread vs Value-Momentum 수익률 ---
    ax2 = axes[1]
    ax2.set_title('Figure 4: TED Spread and Value/Momentum Returns', fontsize=13, fontweight='bold')

    if 'TED Spread' in macro:
        ted = macro['TED Spread'].dropna()
        common = ted.index.intersection(global_val.index).intersection(global_mom.index)

        # 12개월 롤링 수익률
        val_roll = global_val.rolling(12).mean() * 12
        mom_roll = global_mom.rolling(12).mean() * 12

        ax2.plot(val_roll.index, val_roll.values, label='Value (12m rolling)', color='#1f77b4', linewidth=1.2)
        ax2.plot(mom_roll.index, mom_roll.values, label='Momentum (12m rolling)', color='#d62728', linewidth=1.2)
        ax2.set_ylabel('Annualized Return', color='black')
        ax2.legend(loc='upper left', fontsize=10)
        ax2.grid(True, alpha=0.3)

        ax2_twin = ax2.twinx()
        ax2_twin.plot(ted.index, ted.values, label='TED Spread', color='gray', linewidth=1, alpha=0.7)
        ax2_twin.set_ylabel('TED Spread (%)', color='gray')
        ax2_twin.legend(loc='upper right', fontsize=10)

        # 1998 LTCM, 2008 GFC 표시
        for date, label in [('1998-09', 'LTCM'), ('2008-10', 'GFC')]:
            ax2.axvline(pd.Timestamp(date), color='red', linestyle='--', alpha=0.5)
            ax2.text(pd.Timestamp(date), ax2.get_ylim()[1] * 0.9, label, fontsize=9, color='red')

    fig.tight_layout()
    fig.savefig(OUTPUT_DIR / 'figure3_4_macro_risk.png', dpi=150, bbox_inches='tight')
    print(f"\n[Figure 3-4] saved: {OUTPUT_DIR / 'figure3_4_macro_risk.png'}")
    plt.close('all')


# ============================================================
# Main
# ============================================================
def main():
    print("=" * 60)
    print("  Table III-IV: Macro/Liquidity Risk Analysis")
    print("=" * 60)

    print("\n[1] 매크로 데이터 로딩...")
    macro = load_macro_data()

    print("\n[2] 팩터 데이터 로딩...")
    factors_df = load_factors()
    print(f"  Factors: {len(factors_df)} months ({factors_df.index[0].strftime('%Y-%m')} ~ {factors_df.index[-1].strftime('%Y-%m')})")

    print("\n[3] Table III: 매크로 변수 노출도...")
    table_III = replicate_table_III(factors_df, macro)

    print("\n[4] Table IV: 다변량 유동성 분해...")
    replicate_table_IV(factors_df, macro)

    print("\n[5] Figure 3-4: 매크로 리스크 시각화...")
    replicate_figures_3_4(factors_df, macro)

    # 핵심 결과 요약
    print("\n" + "=" * 80)
    print("  핵심 발견 요약 (논문 Section III)")
    print("=" * 80)
    print("""
  1. Value는 유동성 리스크에 음(-)으로 노출 (유동성 위기 시 가치주 하락)
  2. Momentum은 유동성 리스크에 양(+)으로 노출 (유동성 위기 시 모멘텀 전략도 하락)
     → 모멘텀은 "crowded trade" — 유동성 위기 시 모두 같은 방향으로 청산
  3. 이 반대 방향 노출이 Value-Momentum 음의 상관관계를 일부 설명
  4. 하지만 경제적 크기는 작음 — 유동성 리스크만으로 프리미엄 전체를 설명 불가
  5. Value+Momentum 조합은 유동성 리스크에 면역 → abnormal return 유지
""")

    print("  완료.")


if __name__ == '__main__':
    main()
